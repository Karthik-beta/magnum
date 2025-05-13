from rest_framework import generics
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import JSONParser
from django.http.response import JsonResponse
from django.db import connection, OperationalError
from rest_framework.decorators import action
from rest_framework import viewsets, serializers, status
from datetime import timedelta, datetime
from rest_framework.response import Response
from django.db.models import Count
from datetime import datetime, time, date
from django.db.models import Q, Count, Sum, F, FloatField
from django.utils import timezone
from pytz import timezone as pytz_timezone
from django.db.models import Sum, Case, When, IntegerField, Count, Max, F, Window, Subquery, OuterRef, Avg
from django.db.models.functions import TruncHour, ExtractHour


from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from django.views.generic import View
from django.http import HttpResponse
from rest_framework.parsers import FileUploadParser
from rest_framework.views import APIView
import pandas as pd
from pytz import timezone
from openpyxl.utils import get_column_letter
from io import BytesIO



from production import models
from production import serializers

from config import models as config_models



class defaultPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class productionPlanningList(generics.ListCreateAPIView):
    queryset = models.productionPlanning.objects.order_by('-id')
    serializer_class = serializers.productionPlanningSerializer
    pagination_class = defaultPagination

class MostOrderedProducts(generics.RetrieveAPIView):
    def get(self, request, *args, **kwargs):
        total_rows = models.productionPlanning.objects.count()
        product_counts = models.productionPlanning.objects.values('product_id').annotate(count=Count('product_id'))
        most_ordered_products = []

        for entry in product_counts:
            product_id = entry['product_id']
            count = entry['count']
            percentage = (count / total_rows) * 100

            most_ordered_products.append({
                'product_id': product_id,
                'category': product_id,
                'percentage': percentage
            })

        return Response({'most_ordered_products': most_ordered_products})
    
class recentOrders(generics.ListAPIView):
    queryset = models.productionPlanning.objects.order_by('-id')
    serializer_class = serializers.recentOrdersSerializer
    pagination_class = defaultPagination

class productionPlanningEdit(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.productionPlanning.objects.all()
    serializer_class = serializers.productionPlanningSerializer
    lookup_url_kwarg = "id"

class lineMachineConfigList(generics.ListCreateAPIView):
    queryset = models.lineMachineConfig.objects.order_by('-id')
    serializer_class = serializers.lineMachineConfigSerializer
    pagination_class = defaultPagination

class lineMachineConfigEdit(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.lineMachineConfig.objects.all()
    serializer_class = serializers.lineMachineConfigSerializer
    lookup_url_kwarg = "id"



# Database Connection Check Function
def check_database_connection(request):
    try:
        with connection.cursor() as cursor:
            # Try to execute a simple query to check the database connection
            cursor.execute("SELECT 1;")
        return JsonResponse({'message': 'Database Connected.'})
    except OperationalError as e:
        # Handle the database connection error
        error_message = str(e)
        return JsonResponse({'error': 'Database connection error', 'message': error_message}, status=500)


# Open Jobworks from Production Planning
class openJobWorks(generics.ListCreateAPIView):
    # queryset = models.productionPlanning.objects.raw('SELECT id, job_id FROM production_planning WHERE planned_date IS NULL ORDER BY id ASC').no_cache()
    queryset = models.productionPlanning.objects.filter(planned_date__isnull=True).order_by('id')
    serializer_class = serializers.openJobWorkSerializer

class productionPlanById(generics.ListCreateAPIView):
    queryset = models.productionPlanning.objects.all()
    serializer_class = serializers.productionPlanningSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['id', 'job_id']
    



'''Data Processing without consideration of start date and start time latest implementation of the above API'''
class LineMachineSlotConfigViewSet(viewsets.ModelViewSet):
    queryset = models.lineMachineSlotConfig.objects.all()
    serializer_class = serializers.lineMachineSlotConfigSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['id', 'job_id']

    @action(detail=False, methods=['POST'])
    def calculate_schedule(self, request):
        # Deserialize input data
        input_serializer = serializers.ScheduleInputSerializer(data=request.data)
        if input_serializer.is_valid():
            product_id = input_serializer.validated_data['product_id']
            quantity = input_serializer.validated_data['quantity']
            start_date = input_serializer.validated_data['start_date']
            job_id = input_serializer.validated_data['job_id']
            company = input_serializer.validated_data['company']
            plant = input_serializer.validated_data['plant']
            shopfloor = input_serializer.validated_data['shopfloor']
            assembly_line = input_serializer.validated_data['assembly_line']
            machine_id = input_serializer.validated_data['machine_id']
            # start_shift = input_serializer.validated_data['start_shift']
            # start_time_str = input_serializer.validated_data['start_time']
            # # Convert start_time from string to datetime
            # start_time = datetime.strptime(start_time_str, '%H:%M')

            try:
                # Check if a schedule already exists for the given machine_id and start_date
                existing_schedule = models.lineMachineSlotConfig.objects.get(
                    machine_id=machine_id,
                    date=start_date
                )

                # If a schedule already exists, return an error response
                return JsonResponse({'error': 'Schedule already exists for the given machine_id and start_date.'}, status=status.HTTP_400_BAD_REQUEST)

            except models.lineMachineSlotConfig.DoesNotExist:
                # Continue with the schedule calculation and creation
    
        


                # try:
                #     product_recipe = models.productionPlanning.objects.get(job_id=job_id)
                #     product_target = product_recipe.product_target
                # except models.productionPlanning.DoesNotExist:
                #     product_target = None

                # job_id = input_serializer.validated_data['job_id']
                try:
                    job_id_instance = models.productionPlanning.objects.get(pk=job_id)
                except models.productionPlanning.DoesNotExist:
                    return Response({'error': 'Invalid job_id provided.'}, status=status.HTTP_400_BAD_REQUEST)
                except ValueError:
                    return Response({'error': 'Invalid job_id format.'}, status=status.HTTP_400_BAD_REQUEST)

                try:
                    product_recipe = job_id_instance
                    product_target = product_recipe.product_target
                except AttributeError:
                    product_target = None

                if product_target:
                    product_target_seconds = product_target.total_seconds()
                else:
                    # Set a default value or raise an exception as needed
                    product_target_seconds = 0


                # Define the product target (60 seconds per unit)
                product_target = product_target_seconds  # seconds

                # Calculate total hours required to produce the quantity
                total_seconds_required = quantity * product_target
                total_hours_required = round(total_seconds_required / 3600, 2)  # 1 hour = 3600 seconds

                # Calculate the number of days required
                days_required = total_hours_required / 21  # Assuming 21 hours per day

                # Initialize the current date and remaining hours
                current_date = start_date
                remaining_hours = total_hours_required
                planned_hours = 0
                total_planned_production = 0
                balance_production = quantity

                while remaining_hours > 0:
                    # Calculate planned production for the day
                    planned_hours = min(21, remaining_hours)

                    # Divide planned_hours by 2 for shift_a and shift_b
                    shift_a_planned_hours = shift_b_planned_hours = shift_c_planned_hours = planned_hours // 3

                    

                    planned_production = planned_hours * (3600 / product_target)
                    # Calculate balance production decrementally

                    total_planned_production += planned_production

                    balance_production = quantity - total_planned_production
                    

                    # Create a LineMachineSlotConfig object and save it to the database
                    line_config = models.lineMachineSlotConfig(
                        product_id=product_id,
                        date=current_date,
                        planned_hours=planned_hours,
                        planned_production=planned_production,  # Planned production in units
                        remaining_hours=remaining_hours - planned_hours,
                        balance_production=balance_production,  # Balance production in units
                        shift_a=f'06 - 14 ({shift_a_planned_hours})',
                        shift_b=f'14 - 22 ({shift_b_planned_hours})',
                        shift_c=f'22 - 06 ({shift_c_planned_hours})',
                        # shift_c= None,
                        job_id=job_id_instance,
                        company=company,
                        plant=plant,
                        shopfloor=shopfloor,
                        assembly_line=assembly_line,
                        machine_id=machine_id,
                    )
                    line_config.save()

                    # Update current date and remaining hours for the next iteration
                    current_date += timedelta(days=1)
                    remaining_hours -= planned_hours

                return Response({'message': 'Schedule calculated and saved.'}, status=status.HTTP_201_CREATED)
    
        else:
            return Response(input_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

    @action(detail=False, methods=['GET'])
    def get_schedule(self, request):
        # Add your code for handling GET requests here
        # For example, you can retrieve and return schedule data
        schedules = models.lineMachineSlotConfig.objects.all()
        serializer = serializers.lineMachineSlotConfigSerializer(schedules, many=True)
        return Response(serializer.data)
    


class DateBlockAPIView(generics.ListAPIView):
    serializer_class = serializers.lineMachineSlotConfigSerializer  # Use your serializer class


    def get_queryset(self):
        # Calculate the first and last dates from the 'date' field
        first_date = models.lineMachineSlotConfig.objects.earliest('date').date
        last_date = models.lineMachineSlotConfig.objects.latest('date').date

        return [{'first_date': first_date, 'last_date': last_date}]

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        return Response(queryset)






class LineMachineSlotConfigEdit(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.lineMachineSlotConfig.objects.all()
    serializer_class = serializers.lineMachineSlotConfigSerializer
    lookup_url_kwarg = "id"



'''Get View for machineWiseDate Model'''
# class machineWiseDataView(generics.ListCreateAPIView):
#     queryset = models.machineWiseData.objects.order_by("time")
#     serializer_class = serializers.machineWiseDataSerializer
#     filter_backends = [DjangoFilterBackend]
#     filterset_fields = ['id', 'date', 'time', 'machine_id', 'product_target']

# '''Get View for machinewise group by machine id'''
# class machineWiseDataView(generics.ListAPIView):
#     # queryset = models.machineWiseData.objects.all()
#     serializer_class = serializers.machineWiseDataSerializer
    
#     now = datetime.now().time()
    
#     def get_queryset(self):
#         now = datetime.now().time()

#         if now >= datetime.strptime("08:00", "%H:%M").time() and now <= datetime.strptime("20:00", "%H:%M").time():
#             # Current time is between 08:00 and 20:00
#             queryset = models.machineWiseData.objects.filter(
#                 date=datetime.today(),
#                 time__in=[f"{i:02d}:00 - {i+1:02d}:00" for i in range(8, 20)]
#             )
#         else:
#             # Current time is between 20:00 and 08:00
#             today_midnight = datetime.combine(datetime.today(), datetime.min.time())
#             tomorrow_midnight = today_midnight + timedelta(days=1)

#             queryset = models.machineWiseData.objects.filter(
#                 date=datetime.today(),
#                 time__in=[f"{i:02d}:00 - {i+1:02d}:00" for i in range(20, 24)] +
#                          [f"{i:02d}:00 - {i+1:02d}:00" for i in range(0, 8)]
#             ).union(
#                 models.machineWiseData.objects.filter(
#                     date=tomorrow_midnight.date(),
#                     time__in=[f"{i:02d}:00 - {i+1:02d}:00" for i in range(0, 8)]
#                 )
#             )

#         return queryset.order_by('id')

import pytz

''' 2.0 Get View for machinewise group by machine id'''
class machineWiseDataView(generics.ListAPIView):
    serializer_class = serializers.machineWiseDataSerializer

    def get_queryset(self):
        ist_timezone = pytz.timezone('Asia/Kolkata')
        now = datetime.now(ist_timezone)

        if datetime.strptime("06:00", "%H:%M").time() <= now.time() <= datetime.strptime("14:00", "%H:%M").time():
            queryset = models.machineWiseData.objects.filter(
                date=now.date(),
                time__in=[f"{i:02d}:00 - {i+1:02d}:00" for i in range(6, 14)]
            )
        elif datetime.strptime("14:00", "%H:%M").time() <= now.time() <= datetime.strptime("21:59", "%H:%M").time():
            queryset = models.machineWiseData.objects.filter(
                date=now.date(),
                time__in=[f"{i:02d}:00 - {i+1:02d}:00" for i in range(14, 23)] 
            )
        elif datetime.strptime("22:00", "%H:%M").time() <= now.time() <= datetime.strptime("23:59", "%H:%M").time():
            queryset = models.machineWiseData.objects.filter(
                date=now.date(),
                time__in=[f"{i:02d}:00 - {i+1:02d}:00" for i in range(22, 24)] + ["23:00 - 00:00"]
            )
        elif datetime.strptime("00:00", "%H:%M").time() <= now.time() <= datetime.strptime("06:00", "%H:%M").time():
            yesterday = now.date() - timedelta(days=1)
            queryset = models.machineWiseData.objects.filter(
                Q(date=yesterday, time__in=[f"{i:02d}:00 - {i+1:02d}:00" for i in range(20, 24)] + ["23:00 - 00:00"]) |
                Q(date=now.date(), time__in=[f"{i:02d}:00 - {i+1:02d}:00" for i in range(0, 6)])
            )
        else:
            # Handle unexpected cases or provide a default queryset
            # Example: return models.machineWiseData.objects.all()
            pass

        return queryset.order_by('id')




'''Update View for machineWiseDate Model'''
class machineWiseDataUpdate(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.machineWiseData.objects.all()
    serializer_class = serializers.machineWiseDataUpdateSerializer
    lookup_url_kwarg = "id" 




class ProductionPlanningStatsView(generics.ListAPIView):
    queryset = models.productionPlanning.objects.all()
    serializer_class = serializers.productionPlanningSerializer

    def get_stats(self):
        today = date.today()
        planned_count = self.queryset.filter(planned_date__isnull=False, processing_date__isnull=True, completed_date__isnull=True).count()
        order_in_process_count = self.queryset.filter(processing_date__isnull=False, completed_date__isnull=True).count()
        completed_count = self.queryset.exclude(completed_date__isnull=True).count()
        transactions_today_count = self.queryset.filter(Q(assigned_date=today) | Q(planned_date=today) | Q(processing_date=today) | Q(completed_date=today)).count()

        return {
            'planned': planned_count,
            'order_in_process': order_in_process_count,
            'completed': completed_count,
            'transactions_today': transactions_today_count,
        }

    def list(self, request, *args, **kwargs):
        stats = self.get_stats()
        return Response(stats)
    

class lineMachineSlotConfigViewAll(generics.ListAPIView):
    queryset = models.lineMachineSlotConfig.objects.all()
    serializer_class = serializers.lineMachineSlotConfigSerializer



class ExportExcelView(View):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Machine Records"

        headers = ["Employee ID", "Employee Name", "Log Date", "Company", "Shift Status"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        for row_num, record in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=record.employeeid)
            ws.cell(row=row_num, column=2, value=record.employee_name)
            ws.cell(row=row_num, column=3, value=record.logdate)
            ws.cell(row=row_num, column=4, value=record.company)
            ws.cell(row=row_num, column=5, value=record.shift_status)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=attendance_records.xlsx"
        wb.save(response)

        return response

    def get_filtered_queryset(self, request):
        queryset = super().get_queryset()
        search_query = self.request.query_params.get('search', None)
        date_query = self.request.query_params.get('date', None)  

        if date_query:
            try:
                parsed_date = self.parse_date_with_format(date_query, "%Y/%m/%d")  
                next_day = parsed_date + timedelta(days=1)  

                queryset = queryset.filter(
                    logdate__gte=parsed_date,
                    logdate__lt=next_day
                )
            except Exception as e:
                print("Error parsing date:", e)

        if search_query:
            queryset = queryset.filter(
                Q(employee_name__icontains=search_query) |
                Q(employeeid__icontains=search_query)
            )
        return queryset
    

class AssemblyLineWiseDataView(generics.ListAPIView):
    queryset = models.assemblyLineWiseData.objects.all()
    serializer_class = serializers.AssemblyLineWiseDataSerializer

    def list(self, request, *args, **kwargs):
        last_object = self.get_queryset().last()
        serializer = self.get_serializer([last_object], many=True)
        return Response(serializer.data)
    

class AssemblyLineWiseDataUpdate(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.assemblyLineWiseData.objects.all()
    serializer_class = serializers.AssemblyLineWiseDataSerializer
    lookup_url_kwarg = "id"

class soloAssemblyLineDataView(generics.ListAPIView):
    # queryset = models.soloAssemblyLineData.objects.all()
    serializer_class = serializers.soloAssemblyLineDataSerializer

    def get_queryset(self):
        current_datetime = datetime.now()
        current_date = current_datetime.date()
        current_time = current_datetime.time()

        # Determine the shift based on the current time
        if time(6, 0) <= current_time <= time(14, 0):
            shift = 'FS'
        elif time(14, 0) <= current_time <= time(22, 0):
            shift = 'SS'
        else:
            shift = 'TS'

        # Filter the queryset based on the current date and shift
        queryset = models.soloAssemblyLineData.objects.filter(date=current_date, shift=shift)

        # Arrange the queryset by 'stage_no' in ascending order
        queryset = queryset.order_by('stage_no')

        return queryset

class soloAssemblyLineDataUpdate(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.soloAssemblyLineData.objects.all()
    serializer_class = serializers.soloAssemblyLineDataUpdateSerializer
    lookup_url_kwarg = "id"



class spellAssemblyLineDataView(generics.ListAPIView):
    # queryset = models.soloAssemblyLineData.objects.all()
    serializer_class = serializers.spellAssemblyLineDataSerializer

    def get_queryset(self):
        current_datetime = datetime.now()
        current_date = current_datetime.date()
        current_time = current_datetime.time()

        # Determine the shift based on the current time
        if time(6, 0) <= current_time <= time(14, 0):
            shift = 'FS'
        elif time(14, 0) <= current_time <= time(22, 0):
            shift = 'SS'
        else:
            shift = 'TS'

        # Filter the queryset based on the current date and shift
        queryset = models.spellAssemblyLineData.objects.filter(date=current_date, shift=shift)

        # Arrange the queryset by 'stage_no' in ascending order
        queryset = queryset.order_by('stage_no')
        return queryset

class spellAssemblyLineDataUpdate(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.spellAssemblyLineData.objects.all()
    serializer_class = serializers.spellAssemblyLineDataUpdateSerializer
    lookup_url_kwarg = "id"
    







class ProductionAndonView(generics.ListAPIView):

    # queryset = models.ProductionAndon.objects.order_by('-machine_datetime')[:1]
    serializer_class = serializers.ProductionAndonSerializer

    def get_queryset(self):
        # Get the current date and time in the Asia/Kolkata timezone
        today = datetime.now(pytz.timezone('Asia/Kolkata'))

        # Retrieve the latest object from the queryset
        latest_object = models.ProductionAndon.objects.order_by('-machine_datetime').first()

        default_value = [{"r": None }]
        # default_value = models.ProductionAndon.objects.none()

        # Check if the latest object exists
        if latest_object:
            # Get the datetime of the latest object
            latest_datetime = latest_object.machine_datetime

            # Compare today's date and time with the datetime of the latest object
            if today.date() == latest_datetime.date():
                print("Today's date matches the date of the latest object.")
            else:
                print("Today's date does not match the date of the latest object.")

            # Define tolerance in minutes
            tolerance_minutes = 2

            # Calculate the time range within the tolerance
            lower_bound = latest_datetime - timedelta(minutes=tolerance_minutes)
            upper_bound = latest_datetime + timedelta(minutes=tolerance_minutes)

            # Check if the current time falls within the tolerance range
            if lower_bound.time() <= today.time() <= upper_bound.time():
                # Return the queryset
                return models.ProductionAndon.objects.order_by('-machine_datetime')[:1]
            
            else:
                print("The current time is not within {} minutes of the time of the latest object.".format(tolerance_minutes))
                # return default_value
                return [{"error": "No objects found in the queryset."}]

        else:
            # print("No objects found in the queryset.")
            # return default_value
            return [{"error": "No objects found in the queryset."}]

        # return models.ProductionAndon.objects.order_by('-machine_datetime')[:1]

    

from django.db.models import Max
from django.db.models.functions import TruncHour

class HourlyProductionAndon(APIView):
    def get(self, request, *args, **kwargs):
        try:
            # Get the latest 'p' value at the end of each hour
            hourly_data = (
                models.ProductionAndon.objects
                .annotate(hour=TruncHour('machine_datetime'))
                .values('hour')
                .annotate(last_p=Max('p'))
            )

            # Extracting the 'p' value at the end of each hour
            result = [{'hour': entry['hour'], 'last_p': entry['last_p']} for entry in hourly_data]

            return Response(result, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        


class ExportExcelMachineView(View):
    def get(self, request, *args, **kwargs):
        # queryset = models.machineWiseData.objects.all()
        queryset = models.machineWiseData.objects.all().order_by('date', 'time')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Machine Records"

        headers = [
            "Plant", "Shopfloor", "Assembly Line", "Machine ID", "Product ID", "Lot/Batch",
            "Product Target", "Shift", "Date", "Time", "Batch Order Quantity", "Order Processed", "Order Pending", "On Time", "Idle Time", "Idle Reason", "Break",
            "Actual", "Target", "Performance", "Gap", "kW-h"
        ]

        # Set font style and background color for headers
        header_font = Font(size=14, bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        footer_font = Font(bold=True)
        footer_fill = PatternFill(start_color="799184", end_color="799184", fill_type="solid")

        # for col_num, header in enumerate(headers, 1):
        #     cell = ws.cell(row=1, column=col_num, value=header)
        #     cell.font = header_font
        #     cell.fill = header_fill

        # Retrieve the last record from productionPlanning
        last_record = models.productionPlanning.objects.latest('id')

        # Get the quantity from the last record
        quantity = last_record.quantity

        processed_quantity = 0
        remaining_quantity = quantity

        row_num = 1
        current_shift = None
        
        shift_total_on_time = 0
        shift_total_idle_time = 0
        shift_total_actual = 0
        shift_total_target = 0
        shift_total_performance = 0
        performance_count = 0
        shift_average_performance = 0
        shift_total_gap = 0

        shift_total_order_quantity = quantity

        for record in queryset:
            # Extract the hour from the record.time in 24-hour format
            record_hour = int(record.time.split(':')[0])

            # Determine the shift based on the time
            if 6 <= record_hour < 14:
                shift = "Shift A, 06 - 14 (7)"
            elif 14 <= record_hour < 22:
                shift = "Shift B, 14 - 22 (7)"
            else:
                shift = "Shift C, 22 - 06 (7)"

            if shift != current_shift:
                # New shift detected, write the previous shift's total target
                if current_shift:
                    ws.cell(row=row_num, column=1, value=f"{current_shift} => Total")
                    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)  # Merge columns 1 to 4
                    for col_num in range(1, 23):  # Loop through columns 1 to 19
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.font = footer_font
                        cell.fill = footer_fill
                    ws.cell(row=row_num, column=11, value=shift_total_order_quantity)
                    ws.cell(row=row_num, column=12, value=processed_quantity)
                    ws.cell(row=row_num, column=13, value=remaining_quantity)
                    ws.cell(row=row_num, column=14, value=shift_total_on_time)
                    ws.cell(row=row_num, column=15, value=shift_total_idle_time)
                    ws.cell(row=row_num, column=18, value=shift_total_actual)
                    ws.cell(row=row_num, column=19, value=shift_total_target)
                    ws.cell(row=row_num, column=20, value=shift_average_performance)
                    ws.cell(row=row_num, column=21, value=shift_total_gap)
                    ws.cell(row=row_num, column=22, value="")
                    row_num += 1

                    ws.cell(row=row_num, column=1, value="")
                    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=22)
                    row_num += 1

                # Reset the shift target sum for the new shift
                shift_total_on_time = 0
                shift_total_idle_time = 0
                shift_total_actual = 0
                shift_total_target = 0
                shift_total_performance = 0
                performance_count = 0
                shift_average_performance = 0
                shift_total_gap = 0


                current_shift = shift
            
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill

                     # Calculate the width of the column based on the length of the header
                    header_length = len(header)
                    column_width = max(16, header_length + 2)  # Minimum width of 10, adjust as needed
                    ws.column_dimensions[get_column_letter(col_num)].width = column_width
                    
                row_num += 1

            if record.actual is not None:
                processed_quantity += record.actual
            else:
                processed_quantity = processed_quantity

            # Write the record data
            ws.cell(row=row_num, column=1, value=record.plant)
            ws.cell(row=row_num, column=2, value=record.shopfloor)
            ws.cell(row=row_num, column=3, value=record.assembly_line)
            ws.cell(row=row_num, column=4, value=record.machine_id)
            ws.cell(row=row_num, column=5, value="ACG")
            ws.cell(row=row_num, column=6, value="TEST")
            ws.cell(row=row_num, column=7, value=record.product_target)
            ws.cell(row=row_num, column=8, value=shift)
            ws.cell(row=row_num, column=9, value=record.date)
            ws.cell(row=row_num, column=10, value=record.time)
            ws.cell(row=row_num, column=11, value=quantity)
            ws.cell(row=row_num, column=12, value=processed_quantity)
            ws.cell(row=row_num, column=13, value=remaining_quantity)
            ws.cell(row=row_num, column=14, value=record.on_time)
            ws.cell(row=row_num, column=15, value=record.idle_time)
            ws.cell(row=row_num, column=16, value="")
            ws.cell(row=row_num, column=17, value="")
            ws.cell(row=row_num, column=18, value=record.actual)
            ws.cell(row=row_num, column=19, value=record.target)
            ws.cell(row=row_num, column=20, value=record.performance)
            ws.cell(row=row_num, column=21, value=record.gap)
            ws.cell(row=row_num, column=22, value=record.current)

            if record.on_time is not None:
                shift_total_on_time += record.on_time
            # shift_total_on_time += record.on_time
            if record.idle_time is not None:
                shift_total_idle_time += record.idle_time
            if record.actual is not None:
                shift_total_actual += record.actual
            else:
                shift_total_actual = shift_total_actual
            # shift_total_actual += record.actual
            if record.target is not None:
                shift_total_target += record.target
            # shift_total_target += record.target
            performance_count += 1
            shift_average_performance = shift_total_performance / performance_count if performance_count > 0 else 0
            shift_total_gap = shift_total_actual - shift_total_target

            # if record.actual is not None:
            #     processed_quantity += record.actual
            # else:
            #     processed_quantity = processed_quantity
            
            if record.actual is not None:
                remaining_quantity -= record.actual
            else:
                remaining_quantity = remaining_quantity

            row_num += 1

        # Write the total target for the last shift
        ws.cell(row=row_num, column=1, value=f"{current_shift} => Total")
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)  # Merge columns 1 to 4
        for col_num in range(1, 23):  # Loop through columns 1 to 19
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = footer_font
            cell.fill = footer_fill
        ws.cell(row=row_num, column=11, value=shift_total_order_quantity)
        ws.cell(row=row_num, column=12, value=processed_quantity)
        ws.cell(row=row_num, column=13, value=remaining_quantity)
        ws.cell(row=row_num, column=14, value=shift_total_on_time)
        ws.cell(row=row_num, column=15, value=shift_total_idle_time)
        ws.cell(row=row_num, column=18, value=shift_total_actual)
        ws.cell(row=row_num, column=19, value=shift_total_target)
        ws.cell(row=row_num, column=20, value=shift_average_performance)
        ws.cell(row=row_num, column=21, value=shift_total_gap)
        ws.cell(row=row_num, column=22, value="")

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=machine_records.xlsx"
        wb.save(response)

        return response
    
# class AllMachinesDetails(generics.ListAPIView):
#     serializer_class = serializers.ProductionAndonSerializer
#     queryset = models.ProductionAndon.objects.none()  # Prevent assertion error by providing an empty queryset

#     def get_machine_ids(self):
#         # Fetch unique machine IDs from ProductionAndon
#         # return models.ProductionAndon.objects.values_list('machine_id', flat=True).distinct()
#         return models.ProductionAndon.objects.filter(machine_id='1')

#     def get_machine_data(self, machine_id):
#         current_time = datetime.now()        
#         time_offset = timedelta(hours=5, minutes=30)
#         current_minutes = current_time.minute
#         hour_minute = current_time.hour + (current_minutes / 60)
        
#         # Shift calculation logic remains the same
#         if 6.5 <= hour_minute < 14.5:
#             start_hour = 6
#             start_minute = 30
#             base_date = current_time
#         elif 14.5 <= hour_minute < 22.5:
#             start_hour = 14
#             start_minute = 30
#             base_date = current_time
#         else:
#             start_hour = 22
#             start_minute = 30
#             base_date = current_time - timedelta(days=1) if hour_minute < 6.5 else current_time

#         start_time = base_date.replace(
#             hour=start_hour, 
#             minute=start_minute, 
#             second=0, 
#             microsecond=0
#         )

#         end_time = start_time + timedelta(hours=8)

#         queryset = models.ProductionAndon.objects.filter(
#             machine_id=machine_id,
#             machine_datetime__gte=start_time + time_offset,
#             machine_datetime__lt=min(end_time + time_offset, current_time + time_offset)
#         ).order_by('machine_datetime')

#         current_slot = start_time + time_offset
#         previous_slot_data = None

#         # Initialize time variables
#         on_time_seconds = 0
#         idle_time_seconds = 0
#         current_group_start = None

#         # Process data in memory
#         while current_slot < (end_time + time_offset) and current_slot < (current_time + time_offset):
#             slot_end = current_slot + timedelta(hours=1)

#             # Filter data for the current slot using dot notation
#             slot_data = [
#                 record for record in queryset 
#                 if current_slot <= record.machine_datetime < slot_end
#             ]

#             # RESET PER-SLOT COUNTERS
#             slot_on_time = 0  # New per-slot counter
#             current_group_start = None
#             last_r_time = None

#             for record in slot_data:
#                 if record.r == 'R':  # Use dot notation here
#                     if current_group_start is None:
#                         current_group_start = record.machine_datetime
#                     last_r_time = record.machine_datetime
#                 else:
#                     if current_group_start is not None:
#                         group_duration = (last_r_time - current_group_start).total_seconds()
#                         slot_on_time += group_duration
#                         current_group_start = None

#             # Handle remaining group after loop
#             if current_group_start is not None and last_r_time is not None:
#                 group_duration = (last_r_time - current_group_start).total_seconds()
#                 slot_on_time += group_duration

#             # Calculate idle time for the current slot
#             adjusted_current_slot = current_slot - time_offset
#             adjusted_slot_end = slot_end - time_offset

#             if adjusted_current_slot <= current_time < adjusted_slot_end:
#                 elapsed_seconds = (current_time - adjusted_current_slot).total_seconds()
#                 idle_time_seconds_current = max(0, elapsed_seconds - slot_on_time)
#             else:
#                 idle_time_seconds_current = max(0, 3600 - slot_on_time)

#             # Accumulate both times
#             on_time_seconds += slot_on_time  # Add slot's on time to total
#             idle_time_seconds += idle_time_seconds_current

#             # Move to next slot
#             current_slot = slot_end

#         break_time_seconds = 0
#         total_time_seconds = on_time_seconds + idle_time_seconds + break_time_seconds

#         # Avoid division by zero if total_time is zero
#         if total_time_seconds == 0:
#             on_time_percentage = 0.0
#             idle_time_percentage = 0.0
#             break_time_percentage = 0.0
#         else:
#             on_time_percentage = round((on_time_seconds / total_time_seconds) * 100, 2)
#             idle_time_percentage = round((idle_time_seconds / total_time_seconds) * 100, 2)
#             break_time_percentage = round((break_time_seconds / total_time_seconds) * 100, 2)

#         # Get production count using dot notation
#         production_count = queryset.last().p if queryset.exists() else 0

#         response_data = {
#             'machine_id': machine_id,
#             'total_on_time': self.seconds_to_hms(int(on_time_seconds)),
#             'total_idle_time': self.seconds_to_hms(int(idle_time_seconds)),
#             'break_time': '00:00:00',
#             'production_count': production_count,
#             'production_target': 480,
#             'health_status': True,
#             'on_time_percentage': on_time_percentage,
#             'idle_time_percentage': idle_time_percentage,
#             'break_time_percentage': break_time_percentage,
#         }

#         return response_data


#     @staticmethod
#     def seconds_to_hms(seconds):
#         """Convert seconds to HH:MM:SS format."""
#         hours, remainder = divmod(seconds, 3600)
#         minutes, seconds = divmod(remainder, 60)
#         return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"

#     def list(self, request, *args, **kwargs):
#         # Get all unique machine IDs
#         # machine_ids = self.get_machine_ids()
#         machine_ids = ['1']
#         result = []

#         # Loop through each machine ID and gather its data
#         for machine_id in machine_ids:
#             machine_data = self.get_machine_data(machine_id)
#             result.append(machine_data)

#         # Return the list of results for all machines
#         return Response(result)

class AllMachinesDetails(generics.ListAPIView):
    serializer_class = serializers.ProductionAndonSerializer
    queryset = models.ProductionAndon.objects.none()  # Prevent assertion error by providing an empty queryset

    def get_machine_ids(self):
        # Fetch unique machine IDs from ProductionAndon
        # return models.ProductionAndon.objects.values_list('machine_id', flat=True).distinct()
        return models.ProductionAndon.objects.filter(machine_id='1')

    def get_machine_data(self, machine_id):
        current_time = datetime.now()
        time_offset = timedelta(hours=5, minutes=30)
        current_minutes = current_time.minute
        hour_minute = current_time.hour + (current_minutes / 60)

        # Shift calculation logic remains the same
        if 6.5 <= hour_minute < 14.5:
            start_hour = 6
            start_minute = 30
            base_date = current_time
        elif 14.5 <= hour_minute < 22.5:
            start_hour = 14
            start_minute = 30
            base_date = current_time
        else:
            start_hour = 22
            start_minute = 30
            base_date = current_time - timedelta(days=1) if hour_minute < 6.5 else current_time

        start_time = base_date.replace(
            hour=start_hour,
            minute=start_minute,
            second=0,
            microsecond=0
        )

        end_time = start_time + timedelta(hours=8)

        queryset = models.ProductionAndon.objects.filter(
            machine_id=machine_id,
            machine_datetime__gte=start_time + time_offset,
            machine_datetime__lt=min(end_time + time_offset, current_time + time_offset)
        ).order_by('machine_datetime')

        current_slot = start_time + time_offset
        previous_slot_data = None

        # Initialize time variables
        on_time_seconds = 0
        idle_time_seconds = 0
        breakdown_time_seconds = 0

        # Process data in memory
        while current_slot < (end_time + time_offset) and current_slot < (current_time + time_offset):
            slot_end = current_slot + timedelta(hours=1)

            # Filter data for the current slot using dot notation
            slot_data = [
                record for record in queryset
                if current_slot <= record.machine_datetime < slot_end
            ]

            slot_on_time, slot_idle_time, slot_breakdown_time = self.get_idle_and_on_time(slot_data, current_slot, slot_end)

            on_time_seconds += slot_on_time
            idle_time_seconds += slot_idle_time
            breakdown_time_seconds += slot_breakdown_time

            # Move to next slot
            current_slot = slot_end

        break_time_seconds = 0
        total_time_seconds = on_time_seconds + idle_time_seconds + breakdown_time_seconds + break_time_seconds

        # Avoid division by zero if total_time is zero
        if total_time_seconds == 0:
            on_time_percentage = 0.0
            idle_time_percentage = 0.0
            break_time_percentage = 0.0
        else:
            on_time_percentage = round((on_time_seconds / total_time_seconds) * 100, 2)
            idle_time_percentage = round((idle_time_seconds / total_time_seconds) * 100, 2)
            break_time_percentage = round((breakdown_time_seconds / total_time_seconds) * 100, 2)

        # Get production count using dot notation
        production_count = queryset.last().p if queryset.exists() else 0

        response_data = {
            'machine_id': machine_id,
            'total_on_time': self.seconds_to_hms(int(on_time_seconds)),
            'total_idle_time': self.seconds_to_hms(int(idle_time_seconds)),
            'breakdown_time': self.seconds_to_hms(int(breakdown_time_seconds)),
            'production_count': production_count,
            'production_target': 480,
            'health_status': True,
            'on_time_percentage': on_time_percentage,
            'idle_time_percentage': idle_time_percentage,
            'break_time_percentage': break_time_percentage,
        }

        return response_data


    @staticmethod
    def seconds_to_hms(seconds):
        """Convert seconds to HH:MM:SS format."""
        hours, remainder = divmod(seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"

    def list(self, request, *args, **kwargs):
        # Get all unique machine IDs
        # machine_ids = self.get_machine_ids()
        machine_ids = ['1']
        result = []

        # Loop through each machine ID and gather its data
        for machine_id in machine_ids:
            machine_data = self.get_machine_data(machine_id)
            result.append(machine_data)

        # Return the list of results for all machines
        return Response(result)


    def get_idle_and_on_time(self, slot_data, current_slot, slot_end):
        """
        Calculate idle and on time based on gaps between 'R' records.
        Idle time is considered for gaps greater than 1 minute between consecutive 'R' records.
        """
        idle_time_seconds = 0
        on_time_seconds = 0
        breakdown_time_seconds = 0

        time_offset = timedelta(hours=5, minutes=30)

        if not slot_data: # If no data in slot, consider full slot as idle if in past, otherwise elapsed as idle
            adjusted_current_slot = current_slot - time_offset
            adjusted_slot_end = slot_end - time_offset
            if adjusted_slot_end <= datetime.now():
                idle_time_seconds = 3600
            elif adjusted_current_slot <= datetime.now() < adjusted_slot_end:
                idle_time_seconds = max(0, (datetime.now() - adjusted_current_slot).total_seconds())
            return on_time_seconds, idle_time_seconds, breakdown_time_seconds


        r_records = [record for record in slot_data if record.r == 'R']

        if not r_records: # If no 'R' records, consider full slot as idle if in past, otherwise elapsed as idle
            adjusted_current_slot = current_slot - time_offset
            adjusted_slot_end = slot_end - time_offset
            if adjusted_slot_end <= datetime.now():
                idle_time_seconds = 3600
            elif adjusted_current_slot <= datetime.now() < adjusted_slot_end:
                idle_time_seconds = max(0, (datetime.now() - adjusted_current_slot).total_seconds())
            return on_time_seconds, idle_time_seconds, breakdown_time_seconds


        last_record_time = current_slot # Initialize with slot start time to consider gap from slot start to first 'R'

        for record in r_records:
            time_diff = record.machine_datetime - last_record_time
            time_limit = timedelta(minutes=1)
            if time_diff < time_limit:
                # idle_time_seconds += time_diff.total_seconds()
                on_time_seconds += time_diff.total_seconds()
            else:
                on_time_seconds += time_limit.total_seconds()
                breakdown_time_seconds += time_diff.total_seconds() - time_limit.total_seconds()

            last_record_time = record.machine_datetime

        # Calculate total on-time as slot duration minus total idle time (within the slot and before first 'R' if any gap > 1min)
        # on_time_seconds = max(0, 3600 - idle_time_seconds) # Maximum on time is 1 hour slot duration

        adjusted_current_slot = current_slot - time_offset
        adjusted_slot_end = slot_end - time_offset

        if adjusted_slot_end <= datetime.now():
            idle_time_seconds = 3600 - on_time_seconds - breakdown_time_seconds
        elif adjusted_current_slot <= datetime.now() < adjusted_slot_end:
            idle_time_seconds = max(0, (datetime.now() - adjusted_current_slot).total_seconds()) - on_time_seconds - breakdown_time_seconds

        return on_time_seconds, idle_time_seconds, breakdown_time_seconds
    
class SingleMachineDetails(generics.ListAPIView):
    serializer_class = serializers.ProductionAndonSerializer
    queryset = models.ProductionAndon.objects.none()  # Prevent assertion error by providing an empty queryset

    def get_machine_data(self, machine_id):
        current_time = datetime.now()

        # Set time range based on the current time
        if current_time.hour >= 6 and current_time.hour < 14:
            start_time = current_time.replace(hour=6, minute=0, second=0, microsecond=0)
            end_time = current_time.replace(hour=14, minute=0, second=0, microsecond=0)
        elif current_time.hour >= 14 and current_time.hour < 22:
            start_time = current_time.replace(hour=14, minute=0, second=0, microsecond=0)
            end_time = current_time.replace(hour=22, minute=0, second=0, microsecond=0)
        else:
            if current_time.hour >= 22:
                start_time = current_time.replace(hour=22, minute=0, second=0, microsecond=0)
                end_time = (start_time + timedelta(hours=8)).replace(hour=6, minute=0, second=0, microsecond=0)
            else:
                start_time = current_time.replace(hour=22, minute=0, second=0, microsecond=0) - timedelta(days=1)
                end_time = current_time.replace(hour=6, minute=0, second=0, microsecond=0)

        # Query the data for this machine_id within the time range
        queryset = models.ProductionAndon.objects.filter(
            machine_id=machine_id,
            machine_datetime__range=(start_time, end_time)
        )

        # Calculate total on time and idle time in seconds
        total_on_time_seconds = queryset.filter(r='R').count() * 2  # Each 'R' entry represents 2 seconds
        total_idle_time_seconds = queryset.filter(r='I').count() * 2  # Each 'I' entry represents 2 seconds

        # Convert the seconds to HH:MM:SS format
        total_on_time_hms = self.seconds_to_hms(total_on_time_seconds)
        total_idle_time_hms = self.seconds_to_hms(total_idle_time_seconds)

        break_time = timedelta(seconds=2400)  # Fixed break time of 40 minutes
        break_time_seconds = break_time.total_seconds()
        time_difference_seconds = total_on_time_seconds + total_idle_time_seconds + break_time_seconds

        on_time_percentage = round((total_on_time_seconds / time_difference_seconds) * 100, 2)
        idle_time_percentage = round((total_idle_time_seconds / time_difference_seconds) * 100, 2)
        break_time_percentage = round((break_time_seconds / time_difference_seconds) * 100, 2)

        # Prepare the response data for this machine
        response_data = {
            'machine_id': machine_id,
            'total_on_time': total_on_time_hms,
            'total_idle_time': total_idle_time_hms,
            'break_time': '00:40:00',  # Fixed break time
            'production_count': queryset.last().p if queryset.exists() else 0,  # Latest production count
            'production_target': 0,  # Replace with actual target if applicable
            'health_status': True,  # You can customize based on machine performance
            'on_time_percentage': on_time_percentage,
            'idle_time_percentage': idle_time_percentage,
            'break_time_percentage': break_time_percentage,
        }

        result = [response_data]

        return result

    def seconds_to_hms(self, seconds):
        hours, remainder = divmod(seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"

    def list(self, request, *args, **kwargs):
        # Get the machine_id from the query parameters
        machine_id = request.query_params.get('machine_id')

        # Check if machine_id is provided
        if not machine_id:
            return Response({"error": "machine_id parameter is required"}, status=status.HTTP_400_BAD_REQUEST)

        # Get data for the specific machine
        try:
            machine_data = self.get_machine_data(machine_id)
        except models.ProductionAndon.DoesNotExist:
            return Response({"error": "No data found for the given machine_id"}, status=status.HTTP_404_NOT_FOUND)

        # Return the data for the specific machine
        return Response(machine_data)

class ShiftShopfloorStatus(APIView):
    def get_shift_times(self, current_time):
        """Calculate shift start and end times based on current time."""
        current_minutes = current_time.minute
        hour_minute = current_time.hour + (current_minutes / 60)

        if 6.5 <= hour_minute < 14.5:  # First Shift (6:30 AM - 2:30 PM)
            start_hour = 6
            start_minute = 30
            base_date = current_time
        elif 14.5 <= hour_minute < 22.5:  # Second Shift (2:30 PM - 10:30 PM)
            start_hour = 14
            start_minute = 30
            base_date = current_time
        else:  # Night Shift (10:30 PM - 6:30 AM)
            start_hour = 22
            start_minute = 30
            base_date = current_time - timedelta(days=1) if hour_minute < 6.5 else current_time

        start_time = base_date.replace(
            hour=start_hour,
            minute=start_minute,
            second=0,
            microsecond=0
        )
        return start_time, start_time + timedelta(hours=8)

    def get_production_count(self, slot_data, previous_slot_data=None):
        """
        Calculate production count for the current slot.
        Returns absolute count for first hour, differential count for subsequent hours.
        """
        current_count = next(
            (record['p'] for record in reversed(slot_data) if record['machine_id'] == '1'),
            0
        )

        # If this is the first slot or no previous data, return absolute count
        if not previous_slot_data:
            return current_count

        # Get the last count from previous slot
        previous_count = next(
            (record['p'] for record in reversed(previous_slot_data) if record['machine_id'] == '1'),
            0
        )

        # Calculate differential count
        return max(0, current_count - previous_count)

    def get_idle_and_on_time(self, slot_data, current_slot, slot_end):
        """
        Calculate idle and on time based on gaps between 'R' records.
        Idle time is considered for gaps greater than 1 minute between consecutive 'R' records.
        """
        idle_time_seconds = 0
        on_time_seconds = 0
        breakdown_time_seconds = 0

        time_offset = timedelta(hours=5, minutes=30)

        if not slot_data: # If no data in slot, consider full slot as idle if in past, otherwise elapsed as idle
            adjusted_current_slot = current_slot - time_offset
            adjusted_slot_end = slot_end - time_offset
            if adjusted_slot_end <= datetime.now():
                idle_time_seconds = 3600
            elif adjusted_current_slot <= datetime.now() < adjusted_slot_end:
                idle_time_seconds = max(0, (datetime.now() - adjusted_current_slot).total_seconds())
            return on_time_seconds, idle_time_seconds, breakdown_time_seconds


        r_records = [record for record in slot_data if record['r'] == 'R']

        if not r_records: # If no 'R' records, consider full slot as idle if in past, otherwise elapsed as idle
            adjusted_current_slot = current_slot - time_offset
            adjusted_slot_end = slot_end - time_offset
            if adjusted_slot_end <= datetime.now():
                idle_time_seconds = 3600
            elif adjusted_current_slot <= datetime.now() < adjusted_slot_end:
                idle_time_seconds = max(0, (datetime.now() - adjusted_current_slot).total_seconds())
            return on_time_seconds, idle_time_seconds, breakdown_time_seconds


        last_record_time = current_slot # Initialize with slot start time to consider gap from slot start to first 'R'

        for record in r_records:
            time_diff = record['machine_datetime'] - last_record_time
            time_limit = timedelta(minutes=1)
            if time_diff < time_limit:
                # idle_time_seconds += time_diff.total_seconds()
                on_time_seconds += time_diff.total_seconds()
            else:                
                on_time_seconds += time_limit.total_seconds()
                breakdown_time_seconds += time_diff.total_seconds() - time_limit.total_seconds()

            last_record_time = record['machine_datetime']

        # Calculate total on-time as slot duration minus total idle time (within the slot and before first 'R' if any gap > 1min)
        # on_time_seconds = max(0, 3600 - idle_time_seconds) # Maximum on time is 1 hour slot duration

        adjusted_current_slot = current_slot - time_offset
        adjusted_slot_end = slot_end - time_offset

        if adjusted_slot_end <= datetime.now():
            idle_time_seconds = 3600 - on_time_seconds - breakdown_time_seconds
        elif adjusted_current_slot <= datetime.now() < adjusted_slot_end:
            idle_time_seconds = max(0, (datetime.now() - adjusted_current_slot).total_seconds()) - on_time_seconds - breakdown_time_seconds

        return on_time_seconds, idle_time_seconds, breakdown_time_seconds


    def get(self, request):
        current_time = datetime.now()
        start_time, end_time = self.get_shift_times(current_time)

        time_offset = timedelta(hours=5, minutes=30)

        # Fetch all relevant data in a single query
        shift_data = list(
            models.ProductionAndon.objects.filter(
                machine_id = '1',
                machine_datetime__gte=start_time + time_offset,
                machine_datetime__lt=min(end_time + time_offset, current_time + time_offset)
            ).values(
                's_no', 'machine_datetime', 'r', 'p', 'machine_id'
            ).order_by('machine_datetime')
        )

        # Initialize result list
        hourly_data = []
        current_slot = start_time + time_offset
        previous_slot_data = None

        # Process data in memory
        while current_slot < (end_time + time_offset) and current_slot < (current_time + time_offset):
            slot_end = current_slot + timedelta(hours=1)

            # Filter data for the current slot
            slot_data = [
                record for record in shift_data
                if current_slot <= record['machine_datetime'] < slot_end
            ]

            # Calculate on_time and idle_time based on gaps between 'R' records
            on_time_seconds, idle_time_seconds, breakdown_time_seconds = self.get_idle_and_on_time(slot_data, current_slot, slot_end)

            # Calculate production count (differential or absolute)
            production_count = self.get_production_count(slot_data, previous_slot_data)

            time_offset = timedelta(hours=5, minutes=30)
            adjusted_current_slot = current_slot - time_offset
            adjusted_slot_end = slot_end - time_offset

            hourly_data.append({
                'time_range': f"{adjusted_current_slot.strftime('%H:%M')} - {adjusted_slot_end.strftime('%H:%M')}",
                'on_time': self.seconds_to_hms(int(on_time_seconds)),
                'idle_time': self.seconds_to_hms(int(idle_time_seconds)),
                'breakdown_time': self.seconds_to_hms(breakdown_time_seconds),
                'production_count': production_count,
            })

            # Store current slot data for next iteration
            previous_slot_data = slot_data
            current_slot += timedelta(hours=1)

        return Response(hourly_data)


    # @staticmethod
    # def seconds_to_hms(seconds):
    #     """Convert seconds to HH:MM:SS format."""
    #     hours, remainder = divmod(seconds, 3600)
    #     minutes, seconds = divmod(remainder, 60)
    #     return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"
    @staticmethod
    def seconds_to_hms(seconds):
        """Convert seconds to HH:MM format, rounding seconds to the nearest minute."""
        # Round seconds to the nearest minute (30 seconds rounds up)
        rounded_minutes = round(seconds / 60)
        
        # Convert total minutes to hours and minutes
        hours, minutes = divmod(rounded_minutes, 60)
        
        return f"{int(hours):02}:{int(minutes):02}"

    
class ShiftShopfloorStatus2(APIView):
    def get_shift_times(self, current_time):
        """Calculate shift start and end times based on current time."""
        current_minutes = current_time.minute
        hour_minute = current_time.hour + (current_minutes / 60)

        if 6.5 <= hour_minute < 14.5:  # First Shift (6:30 AM - 2:30 PM)
            start_hour = 6
            start_minute = 30
            base_date = current_time
        elif 14.5 <= hour_minute < 22.5:  # Second Shift (2:30 PM - 10:30 PM)
            start_hour = 14
            start_minute = 30
            base_date = current_time
        else:  # Night Shift (10:30 PM - 6:30 AM)
            start_hour = 22
            start_minute = 30
            base_date = current_time - timedelta(days=1) if hour_minute < 6.5 else current_time

        start_time = base_date.replace(
            hour=start_hour,
            minute=start_minute,
            second=0,
            microsecond=0
        )
        return start_time, start_time + timedelta(hours=8)

    def get_production_count(self, slot_data, previous_slot_data=None):
        """
        Calculate production count for the current slot.
        Returns absolute count for first hour, differential count for subsequent hours.
        """
        current_count = next(
            (record['p'] for record in reversed(slot_data) if record['machine_id'] == '1'),
            0
        )

        # If this is the first slot or no previous data, return absolute count
        if not previous_slot_data:
            return current_count

        # Get the last count from previous slot
        previous_count = next(
            (record['p'] for record in reversed(previous_slot_data) if record['machine_id'] == '1'),
            0
        )

        # Calculate differential count
        return max(0, current_count - previous_count)

    def get_idle_and_on_time(self, slot_data, current_slot, slot_end):
        """
        Calculate idle and on time based on gaps between 'R' records.
        Idle time is considered for gaps greater than 1 minute between consecutive 'R' records.
        """
        idle_time_seconds = 0
        on_time_seconds = 0
        breakdown_time_seconds = 0

        time_offset = timedelta(hours=5, minutes=30)

        if not slot_data: # If no data in slot, consider full slot as idle if in past, otherwise elapsed as idle
            adjusted_current_slot = current_slot - time_offset
            adjusted_slot_end = slot_end - time_offset
            if adjusted_slot_end <= datetime.now():
                idle_time_seconds = 3600
            elif adjusted_current_slot <= datetime.now() < adjusted_slot_end:
                idle_time_seconds = max(0, (datetime.now() - adjusted_current_slot).total_seconds())
            return on_time_seconds, idle_time_seconds, breakdown_time_seconds


        r_records = [record for record in slot_data if record['r'] == 'R']

        if not r_records: # If no 'R' records, consider full slot as idle if in past, otherwise elapsed as idle
            adjusted_current_slot = current_slot - time_offset
            adjusted_slot_end = slot_end - time_offset
            if adjusted_slot_end <= datetime.now():
                idle_time_seconds = 3600
            elif adjusted_current_slot <= datetime.now() < adjusted_slot_end:
                idle_time_seconds = max(0, (datetime.now() - adjusted_current_slot).total_seconds())
            return on_time_seconds, idle_time_seconds, breakdown_time_seconds


        last_record_time = current_slot # Initialize with slot start time to consider gap from slot start to first 'R'

        for record in r_records:
            time_diff = record['machine_datetime'] - last_record_time
            time_limit = timedelta(minutes=1)
            if time_diff < time_limit:
                # idle_time_seconds += time_diff.total_seconds()
                on_time_seconds += time_diff.total_seconds()
            else:                
                on_time_seconds += time_limit.total_seconds()
                breakdown_time_seconds += time_diff.total_seconds() - time_limit.total_seconds()

            last_record_time = record['machine_datetime']

        # Calculate total on-time as slot duration minus total idle time (within the slot and before first 'R' if any gap > 1min)
        # on_time_seconds = max(0, 3600 - idle_time_seconds) # Maximum on time is 1 hour slot duration

        adjusted_current_slot = current_slot - time_offset
        adjusted_slot_end = slot_end - time_offset

        if adjusted_slot_end <= datetime.now():
            idle_time_seconds = 3600 - on_time_seconds
        elif adjusted_current_slot <= datetime.now() < adjusted_slot_end:
            idle_time_seconds = max(0, (datetime.now() - adjusted_current_slot).total_seconds()) - on_time_seconds

        return on_time_seconds, idle_time_seconds, breakdown_time_seconds


    def get(self, request):
        current_time = datetime.now()
        start_time, end_time = self.get_shift_times(current_time)

        time_offset = timedelta(hours=5, minutes=30)

        # Fetch all relevant data in a single query
        shift_data = list(
            models.ProductionAndon.objects.filter(
                machine_id = '1',
                machine_datetime__gte=start_time + time_offset,
                machine_datetime__lt=min(end_time + time_offset, current_time + time_offset)
            ).values(
                's_no', 'machine_datetime', 'r', 'p', 'machine_id'
            ).order_by('machine_datetime')
        )

        # Initialize result list
        hourly_data = []
        current_slot = start_time + time_offset
        previous_slot_data = None

        # Process data in memory
        while current_slot < (end_time + time_offset) and current_slot < (current_time + time_offset):
            slot_end = current_slot + timedelta(hours=1)

            # Filter data for the current slot
            slot_data = [
                record for record in shift_data
                if current_slot <= record['machine_datetime'] < slot_end
            ]

            # Calculate on_time and idle_time based on gaps between 'R' records
            on_time_seconds, idle_time_seconds, breakdown_time_seconds = self.get_idle_and_on_time(slot_data, current_slot, slot_end)

            # Calculate production count (differential or absolute)
            production_count = self.get_production_count(slot_data, previous_slot_data)

            time_offset = timedelta(hours=5, minutes=30)
            adjusted_current_slot = current_slot - time_offset
            adjusted_slot_end = slot_end - time_offset

            hourly_data.append({
                'time_range': f"{adjusted_current_slot.strftime('%H:%M')} - {adjusted_slot_end.strftime('%H:%M')}",
                'on_time': self.seconds_to_hms(int(on_time_seconds)),
                'idle_time': self.seconds_to_hms(int(idle_time_seconds)),
                'breakdown_time': self.seconds_to_hms(breakdown_time_seconds),
                'production_count': production_count,
            })

            # Store current slot data for next iteration
            previous_slot_data = slot_data
            current_slot += timedelta(hours=1)

        return Response(hourly_data)


    # @staticmethod
    # def seconds_to_hms(seconds):
    #     """Convert seconds to HH:MM:SS format."""
    #     hours, remainder = divmod(seconds, 3600)
    #     minutes, seconds = divmod(remainder, 60)
    #     return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"
    @staticmethod
    def seconds_to_hms(seconds):
        """Convert seconds to HH:MM format, rounding seconds to the nearest minute."""
        # Round seconds to the nearest minute (30 seconds rounds up)
        rounded_minutes = round(seconds / 60)
        
        # Convert total minutes to hours and minutes
        hours, minutes = divmod(rounded_minutes, 60)
        
        return f"{int(hours):02}:{int(minutes):02}"


# class ShiftShopfloorExcelView(APIView):
#     def get_shift_times(self, report_date, shift_type):
#         """Calculate shift start/end times in LOCAL time."""
#         if shift_type == 'A':  # 6:30 AM - 2:30 PM
#             start_time = report_date.replace(hour=6, minute=30, second=0, microsecond=0)
#             end_time = start_time + timedelta(hours=8)
#         elif shift_type == 'B':  # 2:30 PM - 10:30 PM
#             start_time = report_date.replace(hour=14, minute=30, second=0, microsecond=0)
#             end_time = start_time + timedelta(hours=8)
#         else:  # Night Shift (10:30 PM - 6:30 AM next day)
#             start_time = report_date.replace(hour=22, minute=30, second=0, microsecond=0)
#             end_time = start_time + timedelta(hours=8)
#         return start_time, end_time

#     def get_shift_data(self, report_date, shift_type):
#         """Fetch and process data for a specific shift."""
#         time_offset = timedelta(hours=5, minutes=30)
#         start_time_local, end_time_local = self.get_shift_times(report_date, shift_type)
        
#         # Convert LOCAL times to UTC for query
#         start_time_utc = start_time_local + time_offset
#         end_time_utc = start_time_utc + timedelta(hours=8)
#         current_time_utc = datetime.now()  # Current time in UTC

#         # Fetch data for the shift
#         shift_data = list(
#             models.ProductionAndon.objects.filter(
#                 machine_id='1',
#                 machine_datetime__gte=start_time_utc,
#                 machine_datetime__lt=min(end_time_utc, current_time_utc + time_offset)
#             ).order_by('machine_datetime').values(
#                 's_no', 'machine_datetime', 'r', 'p'
#             )
#         )

#         hourly_data = []
#         current_slot_local = start_time_local + time_offset
#         current_time_local = datetime.now()

#         # Initialize previous slot data
#         previous_slot_data = None

#         while current_slot_local < (end_time_local + time_offset):
#             slot_end_local = current_slot_local + timedelta(hours=1)
            
#             # Convert slot times to UTC for filtering
#             slot_start_utc = current_slot_local
#             slot_end_utc = slot_end_local

#             # Filter records within this hour slot (UTC)
#             slot_data = [
#                 record for record in shift_data
#                 if slot_start_utc <= record['machine_datetime'] < slot_end_utc
#             ]

#             # Initialize default values for future slots
#             on_time = 0
#             idle_time = 0
#             production_count = 0

#             # Only calculate metrics if slot is in the past/present
#             if current_slot_local <= (current_time_local + time_offset):
#                 # Calculate on_time using state transitions
#                 current_group_start = None
#                 last_r_time = None
#                 for record in sorted(slot_data, key=lambda x: x['machine_datetime']):
#                     local_time = record['machine_datetime'] + time_offset
#                     if record['r'] == 'R':
#                         if current_group_start is None:
#                             current_group_start = local_time
#                         last_r_time = local_time
#                     else:
#                         if current_group_start and last_r_time:
#                             on_time += (last_r_time - current_group_start).total_seconds()
#                             current_group_start = None
#                 if current_group_start and last_r_time:
#                     on_time += (last_r_time - current_group_start).total_seconds()

#                 # Calculate production count
#                 production_count = self.get_production_count(slot_data, previous_slot_data)

#                 adjusted_current_slot = current_slot_local - time_offset
#                 adjusted_slot_end = slot_end_local - time_offset

#                 # Calculate idle time
#                 if adjusted_current_slot <= current_time_local < adjusted_slot_end:  # Current hour
#                     elapsed = (current_time_local - adjusted_current_slot).total_seconds()
#                     idle_time = max(0, elapsed - on_time)
#                 else:  # Past hour
#                     idle_time = max(0, 3600 - on_time)

#             adjusted_current_slot = current_slot_local - time_offset
#             hourly_data.append({
#                 'hour': adjusted_current_slot,
#                 'on_time': on_time,
#                 'idle_time': idle_time,
#                 'production_count': production_count
#             })

#             previous_slot_data = slot_data
#             current_slot_local += timedelta(hours=1)

#         return hourly_data

#     # def get_production_count(self, slot_data):
#     #     """Calculate production count for the slot."""
#     #     if not slot_data:
#     #         return 0
#     #     return max(record['p'] for record in slot_data if record['p'] is not None) if slot_data else 0
    
#     def get_production_count(self, slot_data, previous_slot_data=None):
#         """
#         Calculate production count for the current slot.
#         Returns absolute count for the first hour, differential count for subsequent hours.
#         """
#         if not slot_data:
#             return 0

#         # Get the maximum production count for the current slot
#         current_count = max(
#             (record['p'] for record in slot_data if record['p'] is not None),
#             default=0
#         )

#         # If this is the first slot or no previous data, return absolute count
#         if not previous_slot_data:
#             return current_count

#         # Get the maximum production count from the previous slot
#         previous_count = max(
#             (record['p'] for record in previous_slot_data if record['p'] is not None),
#             default=0
#         )

#         # Calculate differential count
#         return max(0, current_count - previous_count)

#     # Keep create_sheet, seconds_to_hms, and get methods unchanged
#     # ... (rest of the code remains the same as original)

#     def create_sheet(self, wb, shift_data, shift_period, report_date):
#         """Create a worksheet for a specific shift."""
#         ws = wb.create_sheet(f'Shift {shift_period}')
        
#         # Define styles
#         header_font = Font(bold=True, color="FFFFFF")
#         header_fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")
#         header_alignment = Alignment(horizontal='center', vertical='center')
#         border = Border(
#             left=Side(style='thin'),
#             right=Side(style='thin'),
#             top=Side(style='thin'),
#             bottom=Side(style='thin')
#         )

#         # Write primary headers
#         primary_headers = ['Shopfloor', 'Assembly Line', 'Date', 'Shift']
#         for col, header in enumerate(primary_headers, 1):
#             cell = ws.cell(row=1, column=col, value=header)
#             cell.font = header_font
#             cell.fill = header_fill
#             cell.alignment = header_alignment
#             cell.border = border

#         # Write primary data
#         ws.cell(row=2, column=1, value='ORT Shopfloor').border = border
#         ws.cell(row=2, column=2, value='Line 1 Spindle Production').border = border
#         ws.cell(row=2, column=3, value=report_date.strftime('%Y-%m-%d')).border = border
#         ws.cell(row=2, column=4, value=f'Shift {shift_period}').border = border

#         # Write secondary headers
#         headers = ['Time Range', 'On Time (HH:MM:SS)', 'Idle Time (HH:MM:SS)', 
#                   'Production Count']
        
#         for col, header in enumerate(headers, 1):
#             cell = ws.cell(row=3, column=col, value=header)
#             cell.font = header_font
#             cell.fill = header_fill
#             cell.alignment = header_alignment
#             cell.border = border

#         # Write data rows
#         row = 4
        
#         for stat in shift_data:
#             hour_start = stat['hour']
#             hour_end = hour_start + timedelta(hours=1)
            
            
#             ws.cell(row=row, column=1, value=f"{hour_start.strftime('%H:%M')} - {hour_end.strftime('%H:%M')}").border = border
#             ws.cell(row=row, column=2, value=self.seconds_to_hms(int(stat['on_time']))).border = border
#             ws.cell(row=row, column=3, value=self.seconds_to_hms(int(stat['idle_time']))).border = border
#             # ws.cell(row=row, column=4, value=production_count).border = border
#             ws.cell(row=row, column=4, value=stat['production_count']).border = border
            
#             row += 1

#         # Add summary row
#         summary_font = Font(bold=True)
#         summary_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
#         # Calculate totals
#         total_on_time = sum(stat['on_time'] for stat in shift_data)
#         total_idle_time = sum(stat['idle_time'] for stat in shift_data)
#         total_production = sum(stat['production_count'] for stat in shift_data)
        
#         # Write summary row
#         ws.cell(row=row, column=1, value='Total').font = summary_font
#         ws.cell(row=row, column=1).fill = summary_fill
#         ws.cell(row=row, column=1).border = border
        
#         ws.cell(row=row, column=2, value=self.seconds_to_hms(int(total_on_time))).font = summary_font
#         ws.cell(row=row, column=2).fill = summary_fill
#         ws.cell(row=row, column=2).border = border
        
#         ws.cell(row=row, column=3, value=self.seconds_to_hms(int(total_idle_time))).font = summary_font
#         ws.cell(row=row, column=3).fill = summary_fill
#         ws.cell(row=row, column=3).border = border
        
#         ws.cell(row=row, column=4, value=total_production).font = summary_font
#         ws.cell(row=row, column=4).fill = summary_fill
#         ws.cell(row=row, column=4).border = border

#         # Adjust column widths
#         ws.column_dimensions['A'].width = 20  # Time Range
#         ws.column_dimensions['B'].width = 25  # On Time
#         ws.column_dimensions['C'].width = 22  # Idle Time
#         ws.column_dimensions['D'].width = 18  # Production Count

#     @staticmethod
#     def seconds_to_hms(seconds):
#         """Convert seconds to HH:MM:SS format."""
#         hours, remainder = divmod(seconds, 3600)
#         minutes, seconds = divmod(remainder, 60)
#         return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"

#     def get(self, request, *args, **kwargs):
#         try:
#             # Get date from query parameters, default to today
#             date_str = request.GET.get('date')
#             if date_str:
#                 report_date = datetime.strptime(date_str, '%Y-%m-%d')
#             else:
#                 report_date = datetime.now()
                
#             # Create workbook
#             wb = openpyxl.Workbook()
            
#             # Define shift times
#             shifts = ['A', 'B', 'C']  # Morning, Afternoon, Night

#             # Create sheet for each shift
#             for shift_type in shifts:
#                 shift_data = self.get_shift_data(report_date, shift_type)
#                 self.create_sheet(wb, shift_data, shift_type, report_date)

#             # Remove default sheet
#             if 'Sheet' in wb.sheetnames:
#                 wb.remove(wb['Sheet'])

#             # Save to BytesIO
#             excel_file = BytesIO()
#             wb.save(excel_file)
#             excel_file.seek(0)
            
#             filename = f"Shift_Report_All_{report_date.strftime('%Y%m%d')}.xlsx"
            
#             response = HttpResponse(
#                 excel_file.read(),
#                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#             )
#             response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
#             return response
            
#         except Exception as e:
#             # print(f"Error generating report: {str(e)}")  # Debug print
#             raise

class ShiftShopfloorExcelView(APIView):
    def get_shift_times(self, report_date, shift_type):
        """Calculate shift start/end times in LOCAL time."""
        if shift_type == 'A':  # 6:30 AM - 2:30 PM
            start_time = report_date.replace(hour=6, minute=30, second=0, microsecond=0)
            end_time = start_time + timedelta(hours=8)
        elif shift_type == 'B':  # 2:30 PM - 10:30 PM
            start_time = report_date.replace(hour=14, minute=30, second=0, microsecond=0)
            end_time = start_time + timedelta(hours=8)
        else:  # Night Shift (10:30 PM - 6:30 AM next day)
            start_time = report_date.replace(hour=22, minute=30, second=0, microsecond=0)
            end_time = start_time + timedelta(hours=8)
        return start_time, end_time
    
    def get_idle_and_on_time(self, slot_data, current_slot, slot_end):
        """
        Calculate idle and on time based on gaps between 'R' records.
        Idle time is considered for gaps greater than 1 minute between consecutive 'R' records.
        """
        idle_time_seconds = 0
        on_time_seconds = 0
        breakdown_time_seconds = 0
        last_r_time = None

        time_offset = timedelta(hours=5, minutes=30)

        if not slot_data: # If no data in slot, consider full slot as idle if in past, otherwise elapsed as idle
            adjusted_current_slot = current_slot - time_offset
            adjusted_slot_end = slot_end - time_offset
            if adjusted_slot_end <= datetime.now():
                idle_time_seconds = 3600
            elif adjusted_current_slot <= datetime.now() < adjusted_slot_end:
                idle_time_seconds = max(0, (datetime.now() - adjusted_current_slot).total_seconds())
            return on_time_seconds, idle_time_seconds, breakdown_time_seconds


        r_records = [record for record in slot_data if record['r'] == 'R']

        if not r_records: # If no 'R' records, consider full slot as idle if in past, otherwise elapsed as idle
            adjusted_current_slot = current_slot - time_offset
            adjusted_slot_end = slot_end - time_offset
            if adjusted_slot_end <= datetime.now():
                idle_time_seconds = 3600
            elif adjusted_current_slot <= datetime.now() < adjusted_slot_end:
                idle_time_seconds = max(0, (datetime.now() - adjusted_current_slot).total_seconds())
            return on_time_seconds, idle_time_seconds, breakdown_time_seconds


        last_record_time = current_slot # Initialize with slot start time to consider gap from slot start to first 'R'

        for record in r_records:
            time_diff = record['machine_datetime'] - last_record_time
            time_limit = timedelta(minutes=1)
            if time_diff < time_limit:
                # idle_time_seconds += time_diff.total_seconds()
                on_time_seconds += time_diff.total_seconds()
            else:                
                on_time_seconds += time_limit.total_seconds()
                breakdown_time_seconds += time_diff.total_seconds() - time_limit.total_seconds()

            last_record_time = record['machine_datetime']

        # Calculate total on-time as slot duration minus total idle time (within the slot and before first 'R' if any gap > 1min)
        # on_time_seconds = max(0, 3600 - idle_time_seconds) # Maximum on time is 1 hour slot duration

        adjusted_current_slot = current_slot - time_offset
        adjusted_slot_end = slot_end - time_offset

        if adjusted_slot_end <= datetime.now():
            idle_time_seconds = 3600 - on_time_seconds - breakdown_time_seconds
        elif adjusted_current_slot <= datetime.now() < adjusted_slot_end:
            idle_time_seconds = max(0, (datetime.now() - adjusted_current_slot).total_seconds()) - on_time_seconds - breakdown_time_seconds

        return on_time_seconds, idle_time_seconds, breakdown_time_seconds

    def get_shift_data(self, report_date, shift_type):
        """Fetch and process data for a specific shift."""
        time_offset = timedelta(hours=5, minutes=30)
        start_time_local, end_time_local = self.get_shift_times(report_date, shift_type)
        
        # Convert LOCAL times to UTC for query
        start_time_utc = start_time_local + time_offset
        end_time_utc = start_time_utc + timedelta(hours=8)
        current_time_utc = datetime.now()  # Current time in UTC

        # Fetch data for the shift
        shift_data = list(
            models.ProductionAndon.objects.filter(
                machine_id='1',
                machine_datetime__gte=start_time_utc,
                machine_datetime__lt=min(end_time_utc, current_time_utc + time_offset)
            ).order_by('machine_datetime').values(
                's_no', 'machine_datetime', 'r', 'p'
            )
        )

        hourly_data = []
        current_slot_local = start_time_local + time_offset
        current_time_local = datetime.now()

        # Initialize previous slot data
        previous_slot_data = None

        while current_slot_local < (end_time_local + time_offset):
            slot_end_local = current_slot_local + timedelta(hours=1)
            
            # Convert slot times to UTC for filtering
            slot_start_utc = current_slot_local
            slot_end_utc = slot_end_local

            # Filter records within this hour slot (UTC)
            slot_data = [
                record for record in shift_data
                if slot_start_utc <= record['machine_datetime'] < slot_end_utc
            ]

            # Initialize default values for future slots
            on_time = 0
            idle_time = 0
            breakdown_time = 0
            production_count = 0

            # Only calculate metrics if slot is in the past/present
            if current_slot_local <= (current_time_local + time_offset):
                # Calculate on_time and idle_time based on gaps between 'R' records
                on_time, idle_time, breakdown_time = self.get_idle_and_on_time(slot_data, current_slot_local, slot_end_local)

                # Calculate production count
                production_count = self.get_production_count(slot_data, previous_slot_data)

                adjusted_current_slot = current_slot_local - time_offset
                adjusted_slot_end = slot_end_local - time_offset

                # Calculate idle time
                # if adjusted_current_slot <= current_time_local < adjusted_slot_end:  # Current hour
                #     elapsed = (current_time_local - adjusted_current_slot).total_seconds()
                #     idle_time = max(0, elapsed - on_time)
                # else:  # Past hour
                #     idle_time = max(0, 3600 - on_time) # this line was commented and replaced by above function call

            adjusted_current_slot = current_slot_local - time_offset
            hourly_data.append({
                'hour': adjusted_current_slot,
                'on_time': on_time,
                'idle_time': idle_time,
                'breakdown_time': breakdown_time,
                'production_count': production_count
            })

            previous_slot_data = slot_data
            current_slot_local += timedelta(hours=1)

        return hourly_data

    # def get_production_count(self, slot_data):
    #     """Calculate production count for the slot."""
    #     if not slot_data:
    #         return 0
    #     return max(record['p'] for record in slot_data if record['p'] is not None) if slot_data else 0
    
    def get_production_count(self, slot_data, previous_slot_data=None):
        """
        Calculate production count for the current slot.
        Returns absolute count for the first hour, differential count for subsequent hours.
        """
        if not slot_data:
            return 0

        # Get the maximum production count for the current slot
        current_count = max(
            (record['p'] for record in slot_data if record['p'] is not None),
            default=0
        )

        # If this is the first slot or no previous data, return absolute count
        if not previous_slot_data:
            return current_count

        # Get the maximum production count from the previous slot
        previous_count = max(
            (record['p'] for record in previous_slot_data if record['p'] is not None),
            default=0
        )

        # Calculate differential count
        return max(0, current_count - previous_count)

    # Keep create_sheet, seconds_to_hms, and get methods unchanged
    # ... (rest of the code remains the same as original)

    def create_sheet(self, wb, shift_data, shift_period, report_date):
        """Create a worksheet for a specific shift."""
        ws = wb.create_sheet(f'Shift {shift_period}')
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write primary headers
        primary_headers = ['Shopfloor', 'Assembly Line', 'Date', 'Shift']
        for col, header in enumerate(primary_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write primary data
        ws.cell(row=2, column=1, value='ORT Shopfloor').border = border
        ws.cell(row=2, column=2, value='Line 1 Spindle Production').border = border
        ws.cell(row=2, column=3, value=report_date.strftime('%Y-%m-%d')).border = border
        ws.cell(row=2, column=4, value=f'Shift {shift_period}').border = border

        # Write secondary headers
        headers = ['Time Range', 'On Time (HH:MM:SS)', 'Idle Time (HH:MM)', 'Breakdown Time (HH:MM)',
                  'Production Count']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write data rows
        row = 4
        
        for stat in shift_data:
            hour_start = stat['hour']
            hour_end = hour_start + timedelta(hours=1)
            
            
            ws.cell(row=row, column=1, value=f"{hour_start.strftime('%H:%M')} - {hour_end.strftime('%H:%M')}").border = border
            ws.cell(row=row, column=2, value=self.seconds_to_hms(int(stat['on_time']))).border = border
            ws.cell(row=row, column=3, value=self.seconds_to_hms(int(stat['idle_time']))).border = border
            ws.cell(row=row, column=4, value=self.seconds_to_hms(int(stat['breakdown_time']))).border = border
            # ws.cell(row=row, column=4, value=production_count).border = border
            ws.cell(row=row, column=5, value=stat['production_count']).border = border
            
            row += 1

        # Add summary row
        summary_font = Font(bold=True)
        summary_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Calculate totals
        total_on_time = sum(stat['on_time'] for stat in shift_data)
        total_idle_time = sum(stat['idle_time'] for stat in shift_data)
        total_breakdown_time = sum(stat['breakdown_time'] for stat in shift_data)
        total_production = sum(stat['production_count'] for stat in shift_data)
        
        # Write summary row
        ws.cell(row=row, column=1, value='Total').font = summary_font
        ws.cell(row=row, column=1).fill = summary_fill
        ws.cell(row=row, column=1).border = border
        
        ws.cell(row=row, column=2, value=self.seconds_to_hms(int(total_on_time))).font = summary_font
        ws.cell(row=row, column=2).fill = summary_fill
        ws.cell(row=row, column=2).border = border
        
        ws.cell(row=row, column=3, value=self.seconds_to_hms(int(total_idle_time))).font = summary_font
        ws.cell(row=row, column=3).fill = summary_fill
        ws.cell(row=row, column=3).border = border

        ws.cell(row=row, column=4, value=self.seconds_to_hms(int(total_breakdown_time))).font = summary_font
        ws.cell(row=row, column=4).fill = summary_fill
        ws.cell(row=row, column=4).border = border
        
        ws.cell(row=row, column=5, value=total_production).font = summary_font
        ws.cell(row=row, column=5).fill = summary_fill
        ws.cell(row=row, column=5).border = border

        # Adjust column widths
        ws.column_dimensions['A'].width = 20  # Time Range
        ws.column_dimensions['B'].width = 25  # On Time
        ws.column_dimensions['C'].width = 22  # Idle Time
        ws.column_dimensions['D'].width = 25  # Breakdown Time
        ws.column_dimensions['E'].width = 22  # Production Count

    # @staticmethod
    # def seconds_to_hms(seconds):
    #     """Convert seconds to HH:MM:SS format."""
    #     hours, remainder = divmod(seconds, 3600)
    #     minutes, seconds = divmod(remainder, 60)
    #     return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"
    @staticmethod
    def seconds_to_hms(seconds):
        """Convert seconds to HH:MM format, rounding seconds to the nearest minute."""
        # Round seconds to the nearest minute (30 seconds rounds up)
        rounded_minutes = round(seconds / 60)
        
        # Convert total minutes to hours and minutes
        hours, minutes = divmod(rounded_minutes, 60)
        
        return f"{int(hours):02}:{int(minutes):02}"


    def get(self, request, *args, **kwargs):
        try:
            # Get date from query parameters, default to today
            date_str = request.GET.get('date')
            if date_str:
                report_date = datetime.strptime(date_str, '%Y-%m-%d')
            else:
                report_date = datetime.now()
                
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Define shift times
            shifts = ['A', 'B', 'C']  # Morning, Afternoon, Night

            # Create sheet for each shift
            for shift_type in shifts:
                shift_data = self.get_shift_data(report_date, shift_type)
                self.create_sheet(wb, shift_data, shift_type, report_date)

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            filename = f"Shift_Report_All_{report_date.strftime('%Y%m%d')}.xlsx"
            
            response = HttpResponse(
                excel_file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            # print(f"Error generating report: {str(e)}")  # Debug print
            raise

class ShiftShopfloorExcelView2(APIView):
    def get_shift_times(self, report_date, shift_type):
        """Calculate shift start/end times in LOCAL time."""
        if shift_type == 'A':  # 6:30 AM - 2:30 PM
            start_time = report_date.replace(hour=6, minute=30, second=0, microsecond=0)
            end_time = start_time + timedelta(hours=8)
        elif shift_type == 'B':  # 2:30 PM - 10:30 PM
            start_time = report_date.replace(hour=14, minute=30, second=0, microsecond=0)
            end_time = start_time + timedelta(hours=8)
        else:  # Night Shift (10:30 PM - 6:30 AM next day)
            start_time = report_date.replace(hour=22, minute=30, second=0, microsecond=0)
            end_time = start_time + timedelta(hours=8)
        return start_time, end_time
    
    def get_idle_and_on_time(self, slot_data, current_slot, slot_end):
        """
        Calculate idle and on time based on gaps between 'R' records.
        Idle time is considered for gaps greater than 1 minute between consecutive 'R' records.
        """
        idle_time_seconds = 0
        on_time_seconds = 0
        breakdown_time_seconds = 0
        last_r_time = None

        time_offset = timedelta(hours=5, minutes=30)

        if not slot_data: # If no data in slot, consider full slot as idle if in past, otherwise elapsed as idle
            adjusted_current_slot = current_slot - time_offset
            adjusted_slot_end = slot_end - time_offset
            if adjusted_slot_end <= datetime.now():
                idle_time_seconds = 3600
            elif adjusted_current_slot <= datetime.now() < adjusted_slot_end:
                idle_time_seconds = max(0, (datetime.now() - adjusted_current_slot).total_seconds())
            return on_time_seconds, idle_time_seconds, breakdown_time_seconds


        r_records = [record for record in slot_data if record['r'] == 'R']

        if not r_records: # If no 'R' records, consider full slot as idle if in past, otherwise elapsed as idle
            adjusted_current_slot = current_slot - time_offset
            adjusted_slot_end = slot_end - time_offset
            if adjusted_slot_end <= datetime.now():
                idle_time_seconds = 3600
            elif adjusted_current_slot <= datetime.now() < adjusted_slot_end:
                idle_time_seconds = max(0, (datetime.now() - adjusted_current_slot).total_seconds())
            return on_time_seconds, idle_time_seconds, breakdown_time_seconds


        last_record_time = current_slot # Initialize with slot start time to consider gap from slot start to first 'R'

        for record in r_records:
            time_diff = record['machine_datetime'] - last_record_time
            time_limit = timedelta(minutes=1)
            if time_diff < time_limit:
                # idle_time_seconds += time_diff.total_seconds()
                on_time_seconds += time_diff.total_seconds()
            else:                
                on_time_seconds += time_limit.total_seconds()
                breakdown_time_seconds += time_diff.total_seconds() - time_limit.total_seconds()

            last_record_time = record['machine_datetime']

        # Calculate total on-time as slot duration minus total idle time (within the slot and before first 'R' if any gap > 1min)
        # on_time_seconds = max(0, 3600 - idle_time_seconds) # Maximum on time is 1 hour slot duration

        adjusted_current_slot = current_slot - time_offset
        adjusted_slot_end = slot_end - time_offset

        if adjusted_slot_end <= datetime.now():
            idle_time_seconds = 3600 - on_time_seconds
        elif adjusted_current_slot <= datetime.now() < adjusted_slot_end:
            idle_time_seconds = max(0, (datetime.now() - adjusted_current_slot).total_seconds()) - on_time_seconds

        return on_time_seconds, idle_time_seconds, breakdown_time_seconds

    def get_shift_data(self, report_date, shift_type):
        """Fetch and process data for a specific shift."""
        time_offset = timedelta(hours=5, minutes=30)
        start_time_local, end_time_local = self.get_shift_times(report_date, shift_type)
        
        # Convert LOCAL times to UTC for query
        start_time_utc = start_time_local + time_offset
        end_time_utc = start_time_utc + timedelta(hours=8)
        current_time_utc = datetime.now()  # Current time in UTC

        # Fetch data for the shift
        shift_data = list(
            models.ProductionAndon.objects.filter(
                machine_id='1',
                machine_datetime__gte=start_time_utc,
                machine_datetime__lt=min(end_time_utc, current_time_utc + time_offset)
            ).order_by('machine_datetime').values(
                's_no', 'machine_datetime', 'r', 'p'
            )
        )

        hourly_data = []
        current_slot_local = start_time_local + time_offset
        current_time_local = datetime.now()

        # Initialize previous slot data
        previous_slot_data = None

        while current_slot_local < (end_time_local + time_offset):
            slot_end_local = current_slot_local + timedelta(hours=1)
            
            # Convert slot times to UTC for filtering
            slot_start_utc = current_slot_local
            slot_end_utc = slot_end_local

            # Filter records within this hour slot (UTC)
            slot_data = [
                record for record in shift_data
                if slot_start_utc <= record['machine_datetime'] < slot_end_utc
            ]

            # Initialize default values for future slots
            on_time = 0
            idle_time = 0
            breakdown_time = 0
            production_count = 0

            # Only calculate metrics if slot is in the past/present
            if current_slot_local <= (current_time_local + time_offset):
                # Calculate on_time and idle_time based on gaps between 'R' records
                on_time, idle_time, breakdown_time = self.get_idle_and_on_time(slot_data, current_slot_local, slot_end_local)

                # Calculate production count
                production_count = self.get_production_count(slot_data, previous_slot_data)

                adjusted_current_slot = current_slot_local - time_offset
                adjusted_slot_end = slot_end_local - time_offset

                # Calculate idle time
                # if adjusted_current_slot <= current_time_local < adjusted_slot_end:  # Current hour
                #     elapsed = (current_time_local - adjusted_current_slot).total_seconds()
                #     idle_time = max(0, elapsed - on_time)
                # else:  # Past hour
                #     idle_time = max(0, 3600 - on_time) # this line was commented and replaced by above function call

            adjusted_current_slot = current_slot_local - time_offset
            hourly_data.append({
                'hour': adjusted_current_slot,
                'on_time': on_time,
                'idle_time': idle_time,
                'breakdown_time': breakdown_time,
                'production_count': production_count
            })

            previous_slot_data = slot_data
            current_slot_local += timedelta(hours=1)

        return hourly_data

    # def get_production_count(self, slot_data):
    #     """Calculate production count for the slot."""
    #     if not slot_data:
    #         return 0
    #     return max(record['p'] for record in slot_data if record['p'] is not None) if slot_data else 0
    
    def get_production_count(self, slot_data, previous_slot_data=None):
        """
        Calculate production count for the current slot.
        Returns absolute count for the first hour, differential count for subsequent hours.
        """
        if not slot_data:
            return 0

        # Get the maximum production count for the current slot
        current_count = max(
            (record['p'] for record in slot_data if record['p'] is not None),
            default=0
        )

        # If this is the first slot or no previous data, return absolute count
        if not previous_slot_data:
            return current_count

        # Get the maximum production count from the previous slot
        previous_count = max(
            (record['p'] for record in previous_slot_data if record['p'] is not None),
            default=0
        )

        # Calculate differential count
        return max(0, current_count - previous_count)

    # Keep create_sheet, seconds_to_hms, and get methods unchanged
    # ... (rest of the code remains the same as original)

    def create_sheet(self, wb, shift_data, shift_period, report_date):
        """Create a worksheet for a specific shift."""
        ws = wb.create_sheet(f'Shift {shift_period}')
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write primary headers
        primary_headers = ['Shopfloor', 'Assembly Line', 'Date', 'Shift']
        for col, header in enumerate(primary_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write primary data
        ws.cell(row=2, column=1, value='ORT Shopfloor').border = border
        ws.cell(row=2, column=2, value='Line 1 Spindle Production').border = border
        ws.cell(row=2, column=3, value=report_date.strftime('%Y-%m-%d')).border = border
        ws.cell(row=2, column=4, value=f'Shift {shift_period}').border = border

        # Write secondary headers
        headers = ['Time Range', 'On Time (HH:MM:SS)', 'Idle Time (HH:MM)', 'Production Count']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write data rows
        row = 4
        
        for stat in shift_data:
            hour_start = stat['hour']
            hour_end = hour_start + timedelta(hours=1)
            
            
            ws.cell(row=row, column=1, value=f"{hour_start.strftime('%H:%M')} - {hour_end.strftime('%H:%M')}").border = border
            ws.cell(row=row, column=2, value=self.seconds_to_hms(int(stat['on_time']))).border = border
            ws.cell(row=row, column=3, value=self.seconds_to_hms(int(stat['idle_time']))).border = border
            ws.cell(row=row, column=4, value=stat['production_count']).border = border
            
            row += 1

        # Add summary row
        summary_font = Font(bold=True)
        summary_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Calculate totals
        total_on_time = sum(stat['on_time'] for stat in shift_data)
        total_idle_time = sum(stat['idle_time'] for stat in shift_data)
        total_production = sum(stat['production_count'] for stat in shift_data)
        
        # Write summary row
        ws.cell(row=row, column=1, value='Total').font = summary_font
        ws.cell(row=row, column=1).fill = summary_fill
        ws.cell(row=row, column=1).border = border
        
        ws.cell(row=row, column=2, value=self.seconds_to_hms(int(total_on_time))).font = summary_font
        ws.cell(row=row, column=2).fill = summary_fill
        ws.cell(row=row, column=2).border = border
        
        ws.cell(row=row, column=3, value=self.seconds_to_hms(int(total_idle_time))).font = summary_font
        ws.cell(row=row, column=3).fill = summary_fill
        ws.cell(row=row, column=3).border = border
        
        ws.cell(row=row, column=4, value=total_production).font = summary_font
        ws.cell(row=row, column=4).fill = summary_fill
        ws.cell(row=row, column=4).border = border

        # Adjust column widths
        ws.column_dimensions['A'].width = 20  # Time Range
        ws.column_dimensions['B'].width = 25  # On Time
        ws.column_dimensions['C'].width = 22  # Idle Time
        ws.column_dimensions['E'].width = 22  # Production Count

    # @staticmethod
    # def seconds_to_hms(seconds):
    #     """Convert seconds to HH:MM:SS format."""
    #     hours, remainder = divmod(seconds, 3600)
    #     minutes, seconds = divmod(remainder, 60)
    #     return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"
    @staticmethod
    def seconds_to_hms(seconds):
        """Convert seconds to HH:MM format, rounding seconds to the nearest minute."""
        # Round seconds to the nearest minute (30 seconds rounds up)
        rounded_minutes = round(seconds / 60)
        
        # Convert total minutes to hours and minutes
        hours, minutes = divmod(rounded_minutes, 60)
        
        return f"{int(hours):02}:{int(minutes):02}"


    def get(self, request, *args, **kwargs):
        try:
            # Get date from query parameters, default to today
            date_str = request.GET.get('date')
            if date_str:
                report_date = datetime.strptime(date_str, '%Y-%m-%d')
            else:
                report_date = datetime.now()
                
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Define shift times
            shifts = ['A', 'B', 'C']  # Morning, Afternoon, Night

            # Create sheet for each shift
            for shift_type in shifts:
                shift_data = self.get_shift_data(report_date, shift_type)
                self.create_sheet(wb, shift_data, shift_type, report_date)

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            filename = f"Shift_Report_All_{report_date.strftime('%Y%m%d')}.xlsx"
            
            response = HttpResponse(
                excel_file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            # print(f"Error generating report: {str(e)}")  # Debug print
            raise
