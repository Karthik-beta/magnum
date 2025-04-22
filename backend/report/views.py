from django.shortcuts import render
from rest_framework import generics
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.pagination import PageNumberPagination
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from django.views.generic import View
from django.http import HttpResponse
from django.db.models import Q

from report import models
from report import serializers

from production import models as production_models




class defaultPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class productionReportList(generics.ListCreateAPIView):
    queryset = models.productionReport.objects.order_by('-id')
    serializer_class = serializers.productionReportSerializer
    pagination_class = defaultPagination



class ExportProductionReportExcelView(View):
    def get(self, request, *args, **kwargs):
        queryset = models.productionReport.objects.all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Production Report"

        headers = ["Job ID", "Customer", "PO No", "Batch No", "Assigned Date", "Expected Date", "Company", "Plant", "Shopfloor", "Assembly Line", "Machine ID", "Product ID", "Date", "Shift", "Planned Hours", "Planned Production", "Remaining Hours", "Balance Production", "Manager"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        for row_num, record in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=record.job_id)
            ws.cell(row=row_num, column=2, value=record.customer)
            ws.cell(row=row_num, column=3, value=record.po_no)
            ws.cell(row=row_num, column=4, value=record.batch_no)
            ws.cell(row=row_num, column=5, value=record.assigned_date)
            ws.cell(row=row_num, column=6, value=record.expected_date)
            ws.cell(row=row_num, column=7, value=record.company)
            ws.cell(row=row_num, column=8, value=record.plant)
            ws.cell(row=row_num, column=9, value=record.shopfloor)
            ws.cell(row=row_num, column=10, value=record.assembly_line)
            ws.cell(row=row_num, column=11, value=record.machine_id)
            ws.cell(row=row_num, column=12, value=record.product_id)
            ws.cell(row=row_num, column=13, value=record.date)
            ws.cell(row=row_num, column=14, value=record.shift_a)
            ws.cell(row=row_num, column=15, value=record.shift_b)
            ws.cell(row=row_num, column=16, value=record.shift_c)
            ws.cell(row=row_num, column=17, value=record.planned_hours)
            ws.cell(row=row_num, column=18, value=record.planned_production)
            ws.cell(row=row_num, column=19, value=record.remaining_hours)
            ws.cell(row=row_num, column=20, value=record.balance_production)
            ws.cell(row=row_num, column=21, value=record.manager)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=Production_Report.xlsx"
        wb.save(response)

        return response
    

class ExportProductionInfoExcel(View):
    def get(self, request, *args, **kwargs):
        # queryset = models.machineWiseData.objects.all()
        queryset = production_models.productionPlanning.objects.all()

        lmc = production_models.lineMachineConfig.objects.all()

        last_record = production_models.productionPlanning.objects.latest('id')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Machine Records"

        headers = [
            "Jobwork", "Job Assigned", "Customer", "Po & Date", "Drawing No", "Plant", "Shopfloor", "Assembly Line", 
            "Machine ID", "Product ID", "Lot/Batch", "Product Target", "Shift", "Date", "Time", "Batch Order Quantity",
            "Order Processed", "Order Pending", "Expected Order", "On Time", "Idle Time", "Idle Reason", "Break", "Actual", "Target",
            "Performance", "Gap", "kW-h"
        ]

        row_num = 1

        # Set font style and background color for headers
        header_font = Font(size=14, bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        footer_font = Font(bold=True)
        footer_fill = PatternFill(start_color="799184", end_color="799184", fill_type="solid")

        # Define color codes
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        green_fill = PatternFill(start_color="2F972F", end_color="2F972F", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

            # Calculate the width of the column based on the length of the header
            header_length = len(header)
            column_width = max(20, header_length + 2)  # Minimum width of 20, adjust as needed
            ws.column_dimensions[get_column_letter(col_num)].width = column_width
                    
        row_num += 1

        
        # Dictionary to store plant, shopfloor, and assemblyline details for each job_id
        job_details = {}

        # assigned_date = production_models.productionPlanning.object('assigned_date')

        # Get the quantity from the last record
        quantity = last_record.quantity

        processed_quantity = 0
        remaining_quantity = quantity

        machineWise_data = production_models.machineWiseData.objects.all()

        # Get the list of unique dates
        unique_dates = machineWise_data.values_list('date', flat=True).distinct()


        # Check if any data exists
        if machineWise_data.exists():
            # Get the first object
            first_object = machineWise_data.first()
            
            # Retrieve the first production day
            first_prod_day = first_object.date
        else:
            # Handle the case where no data exists
            first_prod_day = None

        # Populate job_details dictionary with lineMachineConfig details
        for obj2 in lmc:
            job_id = obj2.job_id
            if job_id not in job_details:
                job_details[job_id] = {}

            print("obj2.prprpr:", job_id)

            job_details[job_id]['plant'] = obj2.plant
            job_details[job_id]['shopfloor'] = obj2.shopfloor
            job_details[job_id]['assemblyline'] = obj2.assembly_line
            job_details[job_id]['machine_id'] = obj2.machine_id
            job_details[job_id]['product_id'] = obj2.product_id
            job_details[job_id]['product_target'] = obj2.product_target

            print("obj2.qwqw:", obj2.plant)

        # ws.cell(row=row_num, column=1, value=models.productionPlanning.job_id)
        shifts = ['A', 'B']
        for date_index, date in enumerate(unique_dates):
            # Get machineWiseData for the current date
            machineWise_data_for_date = machineWise_data.filter(date=date)

            # Initialize a variable to store the count of 'actual' field values
            actual_count = 0
            
            # Loop over machineWiseData for the current date to count 'actual' field values
            for machine_data in machineWise_data_for_date:
                actual_count += machine_data.actual

            # Loop over machineWiseData for the current date to count 'target' field values
            target_count = 0
            for machine_data in machineWise_data_for_date:
                target_count += machine_data.target

            # average performance
            performance = 0
            if target_count > 0:
                performance = round((actual_count / target_count) * 100, 2)

            # Loop over machineWiseData for the current date to count 'gap' field values       
            gap = 0
            for machine_data in machineWise_data_for_date:
                gap += machine_data.gap   

            # Loop over machineWiseData for the current date to count 'idle_time' field values 
            idle_time = 0.0
            on_time = 0.0
            for machine_data in machineWise_data_for_date:
                if machine_data.idle_time is not None:
                    idle_time += machine_data.idle_time

                # Check if the value is not None before adding to on_time
                if machine_data.on_time is not None:
                    on_time += machine_data.on_time

            # Loop over machineWiseData for the current date to count 'on_time' field values 
                
            # Loop over machineWiseData for the current date to count 'current' field values
            current = None
            for machine_data in machineWise_data_for_date:
                # Check if current is None before adding machine_data.current value
                if current is None:
                    current = machine_data.current
                else:
                    current += machine_data.current

            for shift in shifts:
                for obj in queryset:

                    job_id = obj.job_id
                    # print("obj.plant:", obj.plant)
                    # print("obj.shopfloor:", obj.shopfloor)
                    # print("obj.assembly_line:", obj.assembly_line)
                    # print("obj.machine_id:", obj.machine_id)
                    print("obj.product_id:", obj.product_id)
                    print("obj.product_target:", obj.product_target)
                    plant = job_details.get(job_id, {}).get('plant', '')
                    # plant = job_details[job_id]['plant'] if job_id in job_details else ""
                    shopfloor = job_details[job_id]['shopfloor'] if job_id in job_details else ""
                    assemblyline = job_details[job_id]['assemblyline'] if job_id in job_details else ""
                    machine_id = job_details[job_id]['machine_id'] if job_id in job_details else ""
                    product_id = job_details[job_id]['product_id'] if job_id in job_details else ""
                    product_target = job_details[job_id]['product_target'] if job_id in job_details else ""

                    # Get the current date from machineWiseData for the current date
                    current_date = machineWise_data_for_date.first().date

                    ws.cell(row=row_num, column=1, value=obj.job_id)
                    ws.cell(row=row_num, column=2, value=obj.assigned_date)
                    ws.cell(row=row_num, column=3, value=obj.customer)
                    ws.cell(row=row_num, column=4, value=obj.po_no)
                    ws.cell(row=row_num, column=5, value=obj.drawing_no)
                    ws.cell(row=row_num, column=6, value=obj2.plant)
                    ws.cell(row=row_num, column=7, value=obj2.shopfloor)
                    ws.cell(row=row_num, column=8, value=obj2.assembly_line)
                    ws.cell(row=row_num, column=9, value=obj2.machine_id)
                    ws.cell(row=row_num, column=10, value=obj.product_id)
                    ws.cell(row=row_num, column=11, value=obj.batch_no)
                    ws.cell(row=row_num, column=12, value=obj.product_target)
                    ws.cell(row=row_num, column=13, value='Shift A, 08 - 20 (11)' if shift == 'A' else 'Shift B, 20 - 08 (11)')
                    ws.cell(row=row_num, column=14, value=current_date)
                    ws.cell(row=row_num, column=15, value='08 - 20' if shift == 'A' else '20 - 08')
                    ws.cell(row=row_num, column=19, value=obj.expected_date)
                    ws.cell(row=row_num, column=20, value=on_time)
                    ws.cell(row=row_num, column=21, value=idle_time)
                    ws.cell(row=row_num, column=24, value=actual_count)
                    ws.cell(row=row_num, column=25, value=target_count)
                    ws.cell(row=row_num, column=26, value=performance)
                    ws.cell(row=row_num, column=27, value=gap)
                    ws.cell(row=row_num, column=28, value=current)

                    row_num += 1

                # ws.append([
                #     obj.job_id,
                #     obj.assigned_date,
                #     obj.customer,
                #     obj.po_no,
                #     obj.drawing_no,
                #     plant,
                #     shopfloor,
                #     assemblyline,
                #     machine_id,
                #     product_id,
                #     obj.batch_no,
                #     product_target,
                # ])
            
            row_num += 1

            ws.cell(row=row_num, column=11, value="Total")
            ws.cell(row=row_num, column=12, value="")
            ws.cell(row=row_num, column=13, value="")
            ws.cell(row=row_num, column=14, value="")
            ws.cell(row=row_num, column=15, value="")
            ws.cell(row=row_num, column=16, value="")
            ws.cell(row=row_num, column=17, value="")
            ws.cell(row=row_num, column=18, value="")
            # ws.cell(row=row_num, column=19, value="")
            ws.cell(row=row_num, column=19, value='Good on process' if obj.completed_date is None else 'Completed')
            ws.cell(row=row_num, column=20, value=on_time)
            ws.cell(row=row_num, column=21, value=idle_time)
            ws.cell(row=row_num, column=22, value="")
            ws.cell(row=row_num, column=23, value="")
            ws.cell(row=row_num, column=24, value=actual_count)
            ws.cell(row=row_num, column=25, value=target_count)
            ws.cell(row=row_num, column=26, value=performance)
            ws.cell(row=row_num, column=27, value=gap)
            ws.cell(row=row_num, column=28, value=current)
            for col_num in range(11, 29):  # Loop through columns 7 to 28
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = footer_font
                cell.fill = footer_fill

                # Check if the current column is column 19
                if col_num == 19:
                    if obj.completed_date is None:
                        cell.font = footer_font
                        cell.fill = green_fill
                    else:
                        cell.font = footer_font
                        cell.fill = orange_fill

            row_num += 2




        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=production_info.xlsx"
        wb.save(response)

        return response
    



class ExportBatchOrderExcel(View):
    def get(self, request, *args, **kwargs):
        # queryset = models.machineWiseData.objects.all()
        queryset = production_models.productionPlanning.objects.all()

        lmc = production_models.lineMachineConfig.objects.all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Machine Records"

        headers = [            
            "Jobwork", "Job Assigned", "Customer", "Po & Date", "Drawing No", "Product ID", "Lot/Batch", "Batch Order Quantity",
            "Order Processed", "Order Pending", "Expected Order", "Date", "Time", "Status"
        ]

        row_num = 1

        # Set font style and background color for headers
        header_font = Font(size=14, bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        footer_font = Font(bold=True)
        footer_fill = PatternFill(start_color="799184", end_color="799184", fill_type="solid")

        # Define color codes
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

            # Calculate the width of the column based on the length of the header
            header_length = len(header)
            column_width = max(20, header_length + 2)  # Minimum width of 20, adjust as needed
            ws.column_dimensions[get_column_letter(col_num)].width = column_width
                    
        row_num += 1

        
        # Dictionary to store plant, shopfloor, and assemblyline details for each job_id
        job_details = {}

        # assigned_date = production_models.productionPlanning.object('assigned_date')

        # Populate job_details dictionary with lineMachineConfig details
        for obj2 in lmc:
            job_id = obj2.job_id
            if job_id not in job_details:
                job_details[job_id] = {}
            job_details[job_id]['plant'] = obj2.plant
            job_details[job_id]['shopfloor'] = obj2.shopfloor
            job_details[job_id]['assemblyline'] = obj2.assembly_line
            job_details[job_id]['machine_id'] = obj2.machine_id
            job_details[job_id]['product_id'] = obj2.product_id
            job_details[job_id]['product_target'] = obj2.product_target

        # ws.cell(row=row_num, column=1, value=models.productionPlanning.job_id)
        shifts = ['A', 'B']
        for shift in shifts:
            for obj in queryset:

                job_id = obj.job_id
                plant = job_details[job_id]['plant'] if job_id in job_details else ""
                shopfloor = job_details[job_id]['shopfloor'] if job_id in job_details else ""
                assemblyline = job_details[job_id]['assemblyline'] if job_id in job_details else ""
                machine_id = job_details[job_id]['machine_id'] if job_id in job_details else ""
                product_id = job_details[job_id]['product_id'] if job_id in job_details else ""
                product_target = job_details[job_id]['product_target'] if job_id in job_details else ""

                ws.cell(row=row_num, column=1, value=obj.job_id)
                ws.cell(row=row_num, column=2, value=obj.assigned_date)
                ws.cell(row=row_num, column=3, value=obj.customer)
                ws.cell(row=row_num, column=4, value=obj.po_no)
                ws.cell(row=row_num, column=5, value=obj.drawing_no)
                ws.cell(row=row_num, column=6, value=obj.product_id)
                ws.cell(row=row_num, column=7, value=obj.batch_no)
                # Set value for column 13 based on the shift
                ws.cell(row=row_num, column=13, value='08 - 20' if shift == 'A' else '20 - 08')
                ws.cell(row=row_num, column=14, value='Processing' if obj.completed_date is None else 'Completed').fill = orange_fill if obj.completed_date is None else green_fill

                row_num += 1



                # ws.append([
                #     obj.job_id,
                #     obj.assigned_date,
                #     obj.customer,
                #     obj.po_no,
                #     obj.drawing_no,
                #     product_id,
                #     obj.batch_no,
                # ])

        row_num += 2

        ws.cell(row=row_num, column=7, value="Total")
        ws.cell(row=row_num, column=8, value="")
        ws.cell(row=row_num, column=9, value="")
        ws.cell(row=row_num, column=10, value="")
        for col_num in range(7, 11):  # Loop through columns 7 to 10
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.font = footer_font
                        cell.fill = footer_fill



        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=Batch_Order.xlsx"
        wb.save(response)

        return response