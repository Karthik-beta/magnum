from django.shortcuts import render
from rest_framework.parsers import JSONParser
from django.http.response import JsonResponse
from rest_framework.decorators import api_view
from rest_framework import status
from rest_framework import generics
from django.views.decorators.csrf import csrf_exempt
from rest_framework.views import APIView
import openpyxl
from openpyxl.styles import Alignment
from django.http import HttpResponse
from datetime import datetime
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Count, Sum
from django.views.generic import View
from datetime import date
from django.db import connection
from django.db.utils import OperationalError
import jwt, datetime
from datetime import datetime, timedelta
from django.db.models import ExpressionWrapper, F, fields, DateTimeField
from django.utils import timezone
from rest_framework.exceptions import AuthenticationFailed
from django.contrib.auth import authenticate
from rest_framework.response import Response
from django.db.models import Q
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict
from django.db.models.functions import TruncMonth
import calendar


from .models import (BreakdownCategory,
                        Company, Location, Shopfloor, Assemblyline, Machine, Andon, BreakdownHMI, AndonData)
from .serializers import (BreakdownCategorySerializer,
                            CompanySerializer, LocationSerializer, ShopfloorSerializer, AssemblylineSerializer,
                            MachineSerializer, AndonSerializer, BreakdownHmiSerializer, AndonDataSerializer)


class BreakdownCategoryListCreateView(generics.ListCreateAPIView):
    queryset = BreakdownCategory.objects.all()
    serializer_class = BreakdownCategorySerializer

class BreakdownCategoryRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = BreakdownCategory.objects.all()
    serializer_class = BreakdownCategorySerializer
    lookup_url_kwarg = "breakdownCategoryId" 

# class ShiftListCreateView(generics.ListCreateAPIView):
#     queryset = Shift.objects.all()
#     serializer_class = ShiftSerializer

# class CompanyListCreateView(generics.ListCreateAPIView):
#     queryset = Company.objects.all()
#     serializer_class = CompanySerializer

class LocationListCreateView(generics.ListCreateAPIView):
    queryset = Location.objects.all()
    serializer_class = LocationSerializer

class ShopfloorListCreateView(generics.ListCreateAPIView):
    queryset = Shopfloor.objects.all()
    serializer_class = ShopfloorSerializer

class AssemblylineListCreateView(generics.ListCreateAPIView):
    queryset = Assemblyline.objects.all()
    serializer_class = AssemblylineSerializer

class MachineListCreateView(generics.ListCreateAPIView):
    queryset = Machine.objects.all()
    serializer_class = MachineSerializer


@csrf_exempt
def andonapi(request, id=0):
    if request.method == 'GET':
        andons = Andon.objects.all()
        andon_serializer = AndonSerializer(andons, many=True, context={'request': request})
        return JsonResponse(andon_serializer.data, safe=False)
    
    elif request.method == 'POST':
        andon_data = JSONParser().parse(request)
        andon_serializer = AndonSerializer(data=andon_data, context={'request': request})
        if andon_serializer.is_valid():
            andon_serializer.save()
            return JsonResponse("Added Successfully", safe=False)
        return JsonResponse(andon_serializer.errors, status=400)

    
    elif request.method == 'PUT':
        andon_data = JSONParser().parse(request)
        andon = Andon.objects.get(AndonId=andon_data['AndonId'])
        andon_serializer = AndonSerializer(andon, data=andon_data, context={'request': request})
        if andon_serializer.is_valid():
            andon_serializer.save()
            return JsonResponse("Updated Successfully", safe=False)
        return JsonResponse(andon_serializer.errors, status=400)
    
    elif request.method == 'DELETE':
        andon = Andon.objects.get(ticket=id)
        andon.delete()
        return JsonResponse("Deleted Successfully", safe=False)
    






# class DownloadAndonData(APIView):
#     def get(self, request):
#         # Query the EmployeeData model to retrieve the data
#         andon_data = Andon.objects.all()

#         # Create a new workbook and add a worksheet
#         wb = openpyxl.Workbook()
#         ws = wb.active

#         # Add data from the Andon model to the worksheet
#         for data in andon_data:
#             # Convert datetime objects to a timezone-aware format with tzinfo set to None
#             andon_alerts = data.andon_alerts.replace(tzinfo=None) if data.andon_alerts else None
#             andon_acknowledge = data.andon_acknowledge.replace(tzinfo=None) if data.andon_acknowledge else None
#             andon_resolved = data.andon_resolved.replace(tzinfo=None) if data.andon_resolved else None

#         # Add column headers to the worksheet
#         ws.append(['Ticket', 'Company', 'Location', 'Shop Floor', 'Assembly Line', 'Machine ID', 'Category', 'Sub Category', 'Alert Shift', 'Andon Alerts', 'Andon Acknowledge', 'Andon Resolved', 'Total Breakdown'])

#         # Add data from the EmployeeData model to the worksheet
#         for data in andon_data:
#             ws.append([data.ticket, data.company, data.location, data.shopfloor, data.assemblyline, data.machineId, data.category, data.sub_category, data.alert_shift, data.andon_alerts, data.andon_acknowledge, data.andon_resolved, data.total_time])

#         # Save the workbook to a response
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=Andon_Data.xlsx'
#         wb.save(response)

#         return response
    

class AndonDataPagination(PageNumberPagination):
    page_size = 10  # Adjust the number of records per page as needed
    page_size_query_param = 'page_size'
    max_page_size = 100  # Set the maximum page size if desired


# class AndonDataListCreateView(generics.ListCreateAPIView):
#     queryset = Andon.objects.all()
#     serializer_class = AndonSerializer
#     pagination_class = AndonDataPagination 
#     filter_backends = [DjangoFilterBackend]  # Add this line to include the filter backend
#     filterset_fields = ['assemblyline', 'machineId', 'category', 'alert_shift']



class DownloadBreakdownHMIData(APIView):
    def get(self, request):
        # Query the EmployeeData model to retrieve the data
        andon_data = BreakdownHMI.objects.all()

        # Create a new workbook and add a worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Add data from the Andon model to the worksheet
        # for data in andon_data:
        #     # Convert datetime objects to a timezone-aware format with tzinfo set to None
        #     andon_alerts = data.andon_alerts.replace(tzinfo=None) if data.andon_alerts else None
        #     andon_acknowledge = data.andon_acknowledge.replace(tzinfo=None) if data.andon_acknowledge else None
        #     andon_resolved = data.andon_resolved.replace(tzinfo=None) if data.andon_resolved else None

        # Add column headers to the worksheet
        ws.append(['id', 'machine_id', 'channel_id', 'timestamp', 'breakdown_alert', 'alert_value'])

        # Add data from the EmployeeData model to the worksheet
        for data in andon_data:
            ws.append([data.id, data.machine_id, data.channel_id, data.timestamp, data.breakdown_alert, data.alert_value])

        # Save the workbook to a response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Andon_Data.xlsx'
        wb.save(response)

        return response
    




class AndonDataListCreateView(generics.ListCreateAPIView):
    queryset = AndonData.objects.order_by('-id')
    serializer_class = AndonDataSerializer
    pagination_class = AndonDataPagination 
    filter_backends = [DjangoFilterBackend]  # Add this line to include the filter backend
    filterset_fields = ['assemblyline', 'machineId', 'category', 'alert_shift', 'andon_resolved', 'company', 'shopfloor']

    def get_queryset(self):
        queryset = super().get_queryset()
        filters = Q()

        # Extract query parameters
        machineId = self.request.query_params.get('machineId')
        andon_resolved_isnull = self.request.query_params.get('andon_resolved_isnull')
        category = self.request.query_params.get('category')
        assemblyline = self.request.query_params.get('assemblyline')
        alert_shift = self.request.query_params.get('alert_shift')

        # Dynamically build filters
        if machineId:
            filters &= Q(machineId=machineId)
        if andon_resolved_isnull is not None:
            filters &= Q(andon_resolved__isnull=(andon_resolved_isnull.lower() == 'true'))
        if category:
            filters &= Q(category=category)
        if assemblyline:
            filters &= Q(assemblyline=assemblyline)
        if alert_shift:
            filters &= Q(alert_shift=alert_shift)

        # Apply filters to the queryset
        return queryset.filter(filters)

class AndonDataListView(generics.ListAPIView):
    queryset = AndonData.objects.all()
    serializer_class = AndonDataSerializer
    pagination_class = AndonDataPagination 
    filter_backends = [DjangoFilterBackend]  # Add this line to include the filter backend
    filterset_fields = ['assemblyline', 'machineId', 'category', 'alert_shift']

class AndonDataRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = AndonData.objects.all()
    serializer_class = AndonDataSerializer
    lookup_url_kwarg = "id"

# class DownloadAndonData(APIView):
#     def get(self, request):
#         # Query the AndonData model to retrieve the data
#         andon_data = AndonData.objects.all()

#         # Create a new workbook and add a worksheet
#         wb = openpyxl.Workbook()
#         ws = wb.active

#         # Add column headers to the worksheet and apply styles
#         header_row = ['Company', 'Location', 'Shopfloor', 'Assemblyline', 'Machine_id', 'Category', 'Sub Category', 'Alert Shift', 'Andon Alert', 'Andon Acknowledge', 'Andon Resolved', 'Total Breakdown']
#         header_font = openpyxl.styles.Font(bold=True)
#         header_fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
#         for col_num, column_title in enumerate(header_row, 1):
#             cell = ws.cell(row=1, column=col_num, value=column_title)
#             cell.font = header_font
#             cell.fill = header_fill
#             cell.alignment = Alignment(wrap_text=False)  # Prevent text wrapping

#         # Add data from the AndonData model to the worksheet
#         for row_num, data in enumerate(andon_data, 2):  # Start from row 2 (after header)
#             ws.append([data.company, data.location, data.shopfloor, data.assemblyline, data.machineId, data.category, data.sub_category, data.alert_shift, data.andon_alerts, data.andon_acknowledge, data.andon_resolved, data.total_time])

#         # AutoFit column width for all columns
#         for column_cells in ws.columns:
#             max_length = 0
#             column = column_cells[0].column_letter  # Get the column name (e.g., 'A', 'B', ...)
#             for cell in column_cells:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(cell.value)
#                 except:
#                     pass
#             adjusted_width = (max_length + 2)  # Add extra padding
#             ws.column_dimensions[column].width = adjusted_width

#         # Save the workbook to a response
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=Andon_Data.xlsx'
#         wb.save(response)

#         return response
    


class AndonMetricsView(View):
    def get(self, request, *args, **kwargs):
        # Total open alerts
        total_open_alerts = AndonData.objects.filter(andon_resolved__isnull=True).count()
        
        # Total acknowledge alerts
        total_acknowledge_alerts = AndonData.objects.filter(Q(andon_alerts__isnull=False) & Q(andon_acknowledge__isnull=False) & Q(andon_resolved__isnull=True)).count()
        
        # Today open alerts
        today_open_alerts = AndonData.objects.filter(andon_alerts__contains=date.today().strftime("%Y-%m-%d"), andon_resolved__isnull=True).count()
        
        # Total resetting alerts
        total_resetting_alerts = AndonData.objects.filter(category="RESETTING", andon_resolved__isnull=True).count()
        
        # Total engineering alerts
        total_engineering_alerts = AndonData.objects.filter(category="ENGINEERING", andon_resolved__isnull=True).count()
        
        # Total quality alerts
        total_quality_alerts = AndonData.objects.filter(category="QUALITY", andon_resolved__isnull=True).count()
        
        # Total mech maint alerts
        total_mech_maint_alerts = AndonData.objects.filter(category="MECHMAINT", andon_resolved__isnull=True).count()
        
        # Total elect maint alerts
        total_elect_maint_alerts = AndonData.objects.filter(category="ELECTMAINT", andon_resolved__isnull=True).count()

        # Total number of alerts
        total_alerts = AndonData.objects.all().count()

        # Total closed alerts
        total_closed_alerts = AndonData.objects.exclude(andon_resolved__isnull=True).count()
        
        # Create a JSON response with the calculated metrics
        response_data = {
            "total_open_alerts": total_open_alerts,
            "total_acknowledge_alerts": total_acknowledge_alerts,
            "today_open_alerts": today_open_alerts,
            "total_resetting_alerts": total_resetting_alerts,
            "total_engineering_alerts": total_engineering_alerts,
            "total_quality_alerts": total_quality_alerts,
            "total_mech_maint_alerts": total_mech_maint_alerts,
            "total_elect_maint_alerts": total_elect_maint_alerts,
            "total_alerts": total_alerts,
            "total_closed_alerts": total_closed_alerts,
        }
        
        return JsonResponse(response_data)
    


class Shopfloorwise(View):
    def get(self, request):
        # Query the AndonData model to retrieve the desired data
        andon_details = AndonData.objects.filter(andon_resolved__isnull=True).values(
            'shopfloor',
            'machineId',
            'category',
            'andon_alerts'
        )

        # Convert the QuerySet to a list of dictionaries
        andon_details_list = list(andon_details)

        # Return the data as JSON response
        return JsonResponse(andon_details_list, safe=False)
    



def check_database_connection(request):
    try:
        with connection.cursor():
            # Try to execute a simple query to check the database connection
            pass
        return JsonResponse({'message': 'Database is connected.'})
    except OperationalError as e:
        # Handle the database connection error
        error_message = str(e)
        return JsonResponse({'error': 'Database connection error', 'message': error_message}, status=500)
    






class AndonDataOpenListView(generics.ListAPIView):
    serializer_class = AndonDataSerializer

    def get_queryset(self):
        return AndonData.objects.filter(andon_resolved__isnull=True)
    



# class AndonDataOpenListView(generics.ListAPIView):
#     serializer_class = AndonDataSerializer

#     def get_queryset(self):
#         queryset = AndonData.objects.filter(andon_resolved__isnull=True)
        
#         # Calculate the total breakdown time for each record
#         for record in queryset:
#             andon_alert_time_str = record.andon_alert  # Assuming 'andon_alert' is the field with date/time as a string
#             if andon_alert_time_str:
#                 andon_alert_time = datetime.strptime(andon_alert_time_str, "%Y-%m-%d %H:%M:%S")  # Adjust the format as needed
#                 current_time = datetime.now()
#                 total_breakdown_time = current_time - andon_alert_time
#                 record.total_time = str(total_breakdown_time)
        
#         return queryset






class DownloadAndonData(APIView):
    def get(self, request):
        # Extract query parameters
        assemblyline = request.query_params.get('assemblyline')
        machineId = request.query_params.get('machineId')
        category = request.query_params.get('category')
        alert_shift = request.query_params.get('alert_shift')

        # Build the filter conditions
        filters = Q()
        if assemblyline:
            filters &= Q(assemblyline=assemblyline)
        if machineId:
            filters &= Q(machineId=machineId)
        if category:
            filters &= Q(category=category)
        if alert_shift:
            filters &= Q(alert_shift=alert_shift)

        # Apply filtering based on the constructed filters
        filtered_data = AndonData.objects.filter(filters)

        # Create a new workbook and add a worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Add column headers to the worksheet and apply styles
        header_row = ['Ticket ID','Company', 'Location', 'Shopfloor', 'Assemblyline', 'Machine_id', 'Category', 'Alert Shift', 'Andon Alert', 'Andon Acknowledge', 'Andon Resolved', 'Response Time', 'Repair Time', 'Total Breakdown']
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for col_num, column_title in enumerate(header_row, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=False)  # Prevent text wrapping

        # Add data from the filtered queryset to the worksheet
        for row_num, data in enumerate(filtered_data, 2):  # Start from row 2 (after header)
            ws.append([data.id, data.company, data.location, data.shopfloor, data.assemblyline, data.machineId, data.category, data.alert_shift, data.andon_alerts, data.andon_acknowledge, data.andon_resolved, data.response_time, data.repair_time, data.total_time])

        # AutoFit column width for all columns
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter  # Get the column name (e.g., 'A', 'B', ...)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Add extra padding
            ws.column_dimensions[column].width = adjusted_width

        # Save the workbook to a response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Andon_Data.xlsx'
        wb.save(response)

        return response
    

class AndonDataOpenResettingListView(generics.ListAPIView):
    serializer_class = AndonDataSerializer

    def get_queryset(self):
        return AndonData.objects.filter(andon_resolved__isnull=True, category="RESETTING")
    


class AndonDataOpenEngineeringListView(generics.ListAPIView):
    serializer_class = AndonDataSerializer

    def get_queryset(self):
        return AndonData.objects.filter(andon_resolved__isnull=True, category="ENGINEERING")


class AndonDataOpenElectListView(generics.ListAPIView):
    serializer_class = AndonDataSerializer

    def get_queryset(self):
        return AndonData.objects.filter(andon_resolved__isnull=True, category="ELECT MAINT")
    


class AndonDataOpenQualityListView(generics.ListAPIView):
    serializer_class = AndonDataSerializer

    def get_queryset(self):
        return AndonData.objects.filter(andon_resolved__isnull=True, category="QUALITY")
    

class AndonDataOpenMechListView(generics.ListAPIView):
    serializer_class = AndonDataSerializer

    def get_queryset(self):
        return AndonData.objects.filter(andon_resolved__isnull=True, category="MECH MAINT")

class AndonDataCategoryStatsAPIView(APIView):
    """
    API to get count of different 'category' in AndonData for today, this week, and monthly (grouped by month).
    """
    def get(self, request):
        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start = today_start - timedelta(days=today_start.weekday())
        year_start = today_start.replace(month=1, day=1)

        # Helper to aggregate counts
        def get_counts(qs):
            return dict(
                qs.values('category')
                  .annotate(count=Count('category'))
                  .order_by('-count')
                  .values_list('category', 'count')
            )

        # Daily
        qs_today = AndonData.objects.filter(andon_alerts__gte=today_start, andon_alerts__lt=now)
        daily_counts = get_counts(qs_today)

        # Weekly
        qs_week = AndonData.objects.filter(andon_alerts__gte=week_start, andon_alerts__lt=now)
        weekly_counts = get_counts(qs_week)

        # Monthly (grouped by month)
        qs_months = (
            AndonData.objects
            .filter(andon_alerts__gte=year_start, andon_alerts__lt=now)
            .annotate(month=TruncMonth('andon_alerts'))
            .values('month', 'category')
            .annotate(count=Count('category'))
            .order_by('month')
        )
        # Build monthly data
        raw_monthly_data = defaultdict(dict)
        for entry in qs_months:
            month_str = entry['month'].strftime('%b')  # 'Jan', 'Feb', etc.
            category = entry['category']
            count = entry['count']
            raw_monthly_data[month_str][category] = count

        # Ensure all months are present, set to None if no data
        monthly_data = {}
        for i in range(1, 13):
            month_name = calendar.month_abbr[i]
            if month_name in raw_monthly_data and raw_monthly_data[month_name]:
                monthly_data[month_name] = raw_monthly_data[month_name]
            else:
                monthly_data[month_name] = None

        # Running vs breakdown
        running_breakdown = {}
        for i in range(1, 13):
            month_name = calendar.month_abbr[i]
            month_start = now.replace(month=i, day=1, hour=0, minute=0, second=0, microsecond=0)
            if i == 12:
                next_month_start = month_start.replace(year=month_start.year + 1, month=1)
            else:
                next_month_start = month_start.replace(month=i+1)
            unresolved_qs = AndonData.objects.filter(
                andon_alerts__gte=month_start,
                andon_alerts__lt=next_month_start
            ).values('machineId').distinct()

            breakdown_count = unresolved_qs.count()

            running_breakdown[month_name] = {
                "running": 5 - breakdown_count if breakdown_count <= 5 else 0,
                "breakdown": breakdown_count
            }

        data = {
            "Daily": daily_counts,
            "Weekly": weekly_counts,
            "current_month": monthly_data.get(now.strftime('%b'), None),
            "Monthly": monthly_data,
            "running_breakdown": running_breakdown,
        }
        return Response(data)

class MachineBreakdownTodayAPIView(APIView):
    """
    API to output current day's count of machine breakdowns for WS-001 to WS-005,
    including running/breakdown percentage and category breakdown.
    """
    def get(self, request):
        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        machines = [f"WS-00{i}" for i in range(1, 6)]

        # Query all breakdowns for today for the 5 machines
        qs = AndonData.objects.filter(
            machineId__in=machines,
            andon_alerts__gte=today_start,
            andon_alerts__lt=now
        )

        # Count breakdowns per machine and category
        machine_category_counts = defaultdict(lambda: defaultdict(int))
        machine_breakdown_counts = defaultdict(int)
        shift_category_counts = defaultdict(int)
        for record in qs:
            machine = record.machineId
            category = record.category or "Unknown"
            machine_category_counts[machine][category] += 1
            machine_breakdown_counts[machine] += 1
            shift_category_counts[category] += 1

        online = 0
        offline = 0
        result = {
            "assets": {
                "Online": 0,
                "Offline": 0
            }
        }

        for machine in machines:
            breakdown_count = machine_breakdown_counts.get(machine, 0)
            if breakdown_count > 0:
                running = max(0, 100 - breakdown_count)
                offline += 1
            else:
                running = 100
                online += 1

            result[machine] = {
                "asset": {
                    "running": running,
                    "breakdown": 100 - running
                },
                "category": dict(machine_category_counts[machine]) if breakdown_count > 0 else {}
            }

        result["assets"]["Online"] = online
        result["assets"]["Offline"] = offline

        # Add shift summary (example: using all breakdowns for the day)
        shift_breakdown = sum(shift_category_counts.values())
        shift_running = max(0, 100 - shift_breakdown)
        result["shift"] = {
            "efficiency": {
                "Running": shift_running,
                "Breakdown": shift_breakdown
            },
            "category": dict(shift_category_counts)
        }

        return Response(result)